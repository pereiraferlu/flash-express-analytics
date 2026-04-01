"""
╔══════════════════════════════════════════════════════════════════╗
║   FLASH EXPRESS ANALYTICS  —  app.py   v7.0                     ║
╠══════════════════════════════════════════════════════════════════╣
║    pip install streamlit plotly pandas openpyxl xlrd xlsxwriter  ║
║    streamlit run app.py                                          ║
╚══════════════════════════════════════════════════════════════════╝
"""
import io
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import xlsxwriter

# ═══════════════════════════════════════════════════════════════════════════
#  CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Flash Expres - Analytics",
    page_icon="▪",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
html,body,[class*="css"]{ font-family:'Inter',sans-serif!important }
.block-container{ padding-top:1.8rem!important; padding-left:2rem!important;
                  padding-right:2rem!important; padding-bottom:2rem!important; max-width:100%!important }
header[data-testid="stHeader"]{ background:transparent!important; border-bottom:none!important }

/* ── Upload widget en español ── */
[data-testid="stFileUploaderDropzoneInstructions"] div span:first-child::before{
    content:"Arrastrá tu archivo aquí"; display:block; font-weight:600; color:#475569 }
[data-testid="stFileUploaderDropzoneInstructions"] div span:first-child{ display:none }
[data-testid="stFileUploaderDropzoneInstructions"] small::before{
    content:"Límite 200MB por archivo · CSV, XLSX, XLS"; display:block; font-size:12px; color:#94A3B8 }
[data-testid="stFileUploaderDropzoneInstructions"] small{ display:none }
/* Botón upload — ocultar TODO texto nativo y reemplazar con ::after */
[data-testid="stFileUploaderDropzone"] button{
    background:#1E6BB5!important; border:none!important;
    border-radius:8px!important; font-weight:700!important;
    font-size:0!important; color:transparent!important;
    position:relative!important; min-width:130px!important }
[data-testid="stFileUploaderDropzone"] button:hover{ background:#0F4C8A!important }
[data-testid="stFileUploaderDropzone"] button *{
    font-size:0!important; color:transparent!important;
    visibility:hidden!important; width:0!important; overflow:hidden!important }
[data-testid="stFileUploaderDropzone"] button::after{
    content:"Cargar Archivo" !important;
    font-size:14px!important; font-weight:700!important; color:#ffffff!important;
    visibility:visible!important; display:block!important;
    position:absolute!important; top:50%!important; left:50%!important;
    transform:translate(-50%,-50%)!important; white-space:nowrap!important }

/* ── Tab SVG Icons — silhouettes profesionales ── */
.stTabs [data-baseweb="tab"] > div > p {
    display: flex; align-items: center; gap: 7px }
.stTabs [data-baseweb="tab"]:nth-child(1) > div > p::before {
    content:""; display:inline-block; width:16px; height:16px; flex-shrink:0;
    background-color:#64748B;
    -webkit-mask:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M16 11c1.66 0 3-1.34 3-3s-1.34-3-3-3-3 1.34-3 3 1.34 3 3 3zm-8 0c1.66 0 3-1.34 3-3S9.66 5 8 5 5 6.34 5 8s1.34 3 3 3zm0 2c-2.33 0-7 1.17-7 3.5V19h14v-2.5C15 14.17 10.33 13 8 13zm8 0c-.29 0-.62.02-.97.05 1.16.84 1.97 1.97 1.97 3.45V19h6v-2.5c0-2.33-4.67-3.5-7-3.5z'/%3E%3C/svg%3E") center/contain no-repeat;
    mask:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M16 11c1.66 0 3-1.34 3-3s-1.34-3-3-3-3 1.34-3 3 1.34 3 3 3zm-8 0c1.66 0 3-1.34 3-3S9.66 5 8 5 5 6.34 5 8s1.34 3 3 3zm0 2c-2.33 0-7 1.17-7 3.5V19h14v-2.5C15 14.17 10.33 13 8 13zm8 0c-.29 0-.62.02-.97.05 1.16.84 1.97 1.97 1.97 3.45V19h6v-2.5c0-2.33-4.67-3.5-7-3.5z'/%3E%3C/svg%3E") center/contain no-repeat }
.stTabs [data-baseweb="tab"]:nth-child(2) > div > p::before {
    content:""; display:inline-block; width:16px; height:16px; flex-shrink:0;
    background-color:#64748B;
    -webkit-mask:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M20 8h-3V4H3c-1.1 0-2 .9-2 2v11h2c0 1.66 1.34 3 3 3s3-1.34 3-3h6c0 1.66 1.34 3 3 3s3-1.34 3-3h2v-5l-3-4zM6 18.5c-.83 0-1.5-.67-1.5-1.5s.67-1.5 1.5-1.5 1.5.67 1.5 1.5-.67 1.5-1.5 1.5zm13.5-9l1.96 2.5H17V9.5h2.5zm-1.5 9c-.83 0-1.5-.67-1.5-1.5s.67-1.5 1.5-1.5 1.5.67 1.5 1.5-.67 1.5-1.5 1.5z'/%3E%3C/svg%3E") center/contain no-repeat;
    mask:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M20 8h-3V4H3c-1.1 0-2 .9-2 2v11h2c0 1.66 1.34 3 3 3s3-1.34 3-3h6c0 1.66 1.34 3 3 3s3-1.34 3-3h2v-5l-3-4zM6 18.5c-.83 0-1.5-.67-1.5-1.5s.67-1.5 1.5-1.5 1.5.67 1.5 1.5-.67 1.5-1.5 1.5zm13.5-9l1.96 2.5H17V9.5h2.5zm-1.5 9c-.83 0-1.5-.67-1.5-1.5s.67-1.5 1.5-1.5 1.5.67 1.5 1.5-.67 1.5-1.5 1.5z'/%3E%3C/svg%3E") center/contain no-repeat }
.stTabs [data-baseweb="tab"]:nth-child(3) > div > p::before {
    content:""; display:inline-block; width:16px; height:16px; flex-shrink:0;
    background-color:#64748B;
    -webkit-mask:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M20 3h-1V1h-2v2H7V1H5v2H4c-1.1 0-2 .9-2 2v16c0 1.1.9 2 2 2h16c1.1 0 2-.9 2-2V5c0-1.1-.9-2-2-2zm0 18H4V8h16v13z'/%3E%3C/svg%3E") center/contain no-repeat;
    mask:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M20 3h-1V1h-2v2H7V1H5v2H4c-1.1 0-2 .9-2 2v16c0 1.1.9 2 2 2h16c1.1 0 2-.9 2-2V5c0-1.1-.9-2-2-2zm0 18H4V8h16v13z'/%3E%3C/svg%3E") center/contain no-repeat }
.stTabs [data-baseweb="tab"]:nth-child(4) > div > p::before {
    content:""; display:inline-block; width:16px; height:16px; flex-shrink:0;
    background-color:#64748B;
    -webkit-mask:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M20 6h-2.18c.07-.44.18-.88.18-1.34C18 2.01 15.99 0 13.34 0c-1.3 0-2.51.52-3.38 1.44L9 3 7.04 1.44C6.17.52 4.96 0 3.66 0 1.01 0-1 2.01-1 4.66c0 .46.09.9.18 1.34H-1c-1.1 0-1.99.9-1.99 2L-3 20c0 1.1.89 2 1.99 2H20c1.1 0 2-.9 2-2V8c0-1.1-.9-2-2-2zm-7.34-4c.97 0 1.68.78 1.68 1.66S13.63 5.32 12.66 5.32s-1.68-.78-1.68-1.66S11.69 2 12.66 2zM4.34 2c.97 0 1.68.78 1.68 1.66S5.31 5.32 4.34 5.32 2.66 4.54 2.66 3.66 3.37 2 4.34 2zM20 20H2V8h18v12z'/%3E%3C/svg%3E") center/contain no-repeat;
    mask:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M20 6h-2.18c.07-.44.18-.88.18-1.34C18 2.01 15.99 0 13.34 0c-1.3 0-2.51.52-3.38 1.44L9 3 7.04 1.44C6.17.52 4.96 0 3.66 0 1.01 0-1 2.01-1 4.66c0 .46.09.9.18 1.34H-1c-1.1 0-1.99.9-1.99 2L-3 20c0 1.1.89 2 1.99 2H20c1.1 0 2-.9 2-2V8c0-1.1-.9-2-2-2zm-7.34-4c.97 0 1.68.78 1.68 1.66S13.63 5.32 12.66 5.32s-1.68-.78-1.68-1.66S11.69 2 12.66 2zM4.34 2c.97 0 1.68.78 1.68 1.66S5.31 5.32 4.34 5.32 2.66 4.54 2.66 3.66 3.37 2 4.34 2zM20 20H2V8h18v12z'/%3E%3C/svg%3E") center/contain no-repeat }
.stTabs [data-baseweb="tab"]:nth-child(5) > div > p::before {
    content:""; display:inline-block; width:16px; height:16px; flex-shrink:0;
    background-color:#64748B;
    -webkit-mask:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M12 2C8.13 2 5 5.13 5 9c0 5.25 7 13 7 13s7-7.75 7-13c0-3.87-3.13-7-7-7zm0 9.5c-1.38 0-2.5-1.12-2.5-2.5s1.12-2.5 2.5-2.5 2.5 1.12 2.5 2.5-1.12 2.5-2.5 2.5z'/%3E%3C/svg%3E") center/contain no-repeat;
    mask:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M12 2C8.13 2 5 5.13 5 9c0 5.25 7 13 7 13s7-7.75 7-13c0-3.87-3.13-7-7-7zm0 9.5c-1.38 0-2.5-1.12-2.5-2.5s1.12-2.5 2.5-2.5 2.5 1.12 2.5 2.5-1.12 2.5-2.5 2.5z'/%3E%3C/svg%3E") center/contain no-repeat }
/* Active tab: icon becomes white */
.stTabs [aria-selected="true"] > div > p::before { background-color:#fff!important }
/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"]{
    gap:4px; background:#EEF2F7; border-radius:12px; padding:4px; margin-bottom:.5rem }
.stTabs [data-baseweb="tab"]{
    border-radius:9px; font-size:13px; font-weight:600;
    color:#64748B!important; padding:8px 20px }
.stTabs [aria-selected="true"]{
    background:#1E6BB5!important; color:#fff!important;
    box-shadow:0 2px 8px rgba(30,107,181,.3) }
/* ── Reemplazar ⬡ con iconos SVG profesionales via CSS content ── */
.stTabs [data-baseweb="tab"]:nth-child(1) [data-testid="stMarkdownContainer"] p::before,
.stTabs [data-baseweb="tab"]:nth-child(1) p::before {
    content: "";
    display:inline-block; width:15px; height:15px; margin-right:6px;
    vertical-align:middle;
    background-color: currentColor;
    -webkit-mask: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M16 11c1.66 0 2.99-1.34 2.99-3S17.66 5 16 5c-1.66 0-3 1.34-3 3s1.34 3 3 3zm-8 0c1.66 0 2.99-1.34 2.99-3S9.66 5 8 5C6.34 5 5 6.34 5 8s1.34 3 3 3zm0 2c-2.33 0-7 1.17-7 3.5V19h14v-2.5c0-2.33-4.67-3.5-7-3.5zm8 0c-.29 0-.62.02-.97.05 1.16.84 1.97 1.97 1.97 3.45V19h6v-2.5c0-2.33-4.67-3.5-7-3.5z'/%3E%3C/svg%3E") center/contain no-repeat;
    mask: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M16 11c1.66 0 2.99-1.34 2.99-3S17.66 5 16 5c-1.66 0-3 1.34-3 3s1.34 3 3 3zm-8 0c1.66 0 2.99-1.34 2.99-3S9.66 5 8 5C6.34 5 5 6.34 5 8s1.34 3 3 3zm0 2c-2.33 0-7 1.17-7 3.5V19h14v-2.5c0-2.33-4.67-3.5-7-3.5zm8 0c-.29 0-.62.02-.97.05 1.16.84 1.97 1.97 1.97 3.45V19h6v-2.5c0-2.33-4.67-3.5-7-3.5z'/%3E%3C/svg%3E") center/contain no-repeat;
}

/* ── KPI card ── */
.kc{ background:#F0F5FB; border:1px solid #D6E4F5; border-radius:12px;
    padding:14px 16px; box-shadow:0 1px 4px rgba(15,28,46,.06);
    position:relative; overflow:hidden;
    height:110px; min-height:110px; max-height:110px;
    display:flex; flex-direction:column; justify-content:center;
    transition:background .18s, box-shadow .18s, transform .15s; cursor:default; }
.kc:hover{ background:#DCE9F5!important; box-shadow:0 4px 16px rgba(30,107,181,.18)!important; transform:translateY(-2px); }
.kc::before{ content:''; position:absolute; top:0; left:0; right:0; height:3px; background:var(--ac,#1E6BB5); border-radius:12px 12px 0 0 }
.kc.g::before{ --ac:#10B981 } .kc.g{ background:#EDF7F2; border-color:#BDE8D7 } .kc.g:hover{ background:#D3EFE3!important }
.kc.a::before{ --ac:#F59E0B } .kc.a{ background:#FDF6EC; border-color:#F5DBA5 } .kc.a:hover{ background:#F5E8C8!important }
.kc.p::before{ --ac:#8B5CF6 } .kc.p{ background:#F3F0FD; border-color:#C9BAF5 } .kc.p:hover{ background:#E4DCFA!important }
.kc.r::before{ --ac:#EF4444 } .kc.r{ background:#FEF0F0; border-color:#F5B8B8 } .kc.r:hover{ background:#FAD9D9!important }
.kc.t::before{ --ac:#06B6D4 } .kc.t{ background:#ECFAFE; border-color:#A9E4EF } .kc.t:hover{ background:#D0F2F9!important }
.kl{ font-size:10px;font-weight:700;color:#64748B;text-transform:uppercase;letter-spacing:1px;margin-bottom:4px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis }
.kv{ font-size:26px;font-weight:800;color:#0F1C2E;line-height:1;margin-bottom:3px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis }
.kv.sm{ font-size:15px!important;font-weight:700!important;white-space:normal;line-height:1.2 }
.ks{ font-size:11px;color:#64748B;white-space:nowrap;overflow:hidden;text-overflow:ellipsis }
.ks.up{ color:#10B981;font-weight:600 } .ks.dn{ color:#EF4444;font-weight:600 }

/* ── Chart card ── */
.cw{ background:#fff;border:1px solid #E2E8F0;border-radius:12px;padding:16px 16px 6px;
     box-shadow:0 1px 4px rgba(15,28,46,.06);margin-bottom:.8rem }
.ct{ font-size:12px;font-weight:700;color:#0F1C2E;text-transform:uppercase;letter-spacing:.5px;margin-bottom:2px }
.cs{ font-size:10px;color:#94A3B8;margin-bottom:8px }

/* ── Expander ── */
[data-testid="stExpander"]{ border-top:1px solid #E8EEF4!important;
    border-left:none!important;border-right:none!important;border-bottom:none!important;
    border-radius:0!important;background:transparent!important }

/* ── Buttons ── */
div[data-testid="stDownloadButton"]>button{
    background:#1E6BB5!important;color:#fff!important;border:none!important;
    border-radius:8px!important;font-weight:700!important;font-size:13px!important;
    padding:0!important;height:38px!important;width:100%!important }
div[data-testid="stDownloadButton"]>button:hover{ background:#0F4C8A!important }
div[data-testid="stButton"]>button{
    height:38px!important;border-radius:8px!important;font-weight:600!important;
    font-size:13px!important;width:100%!important }
/* ── Popup button styles handled inside render_popup CSS injection ── */

::-webkit-scrollbar{ width:5px;height:5px }
::-webkit-scrollbar-track{ background:#F8FAFC }
::-webkit-scrollbar-thumb{ background:#CBD5E1;border-radius:3px }

/* ── Ocultar "Press Enter to apply" en number_input ── */
div[data-testid="InputInstructions"] { display: none !important; }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════
#  CONSTANTES — columnas del Excel de origen
# ═══════════════════════════════════════════════════════════════════════════
COL_PIEZA   = "Pieza"
COL_CLIENTE = "Cliente"
COL_HDR     = "Hoja de Ruta"
COL_ESTADO  = "Estado"
COL_DISTRIB = "Distribuidor"
COL_LOC     = "Localidad Destino"
COL_BULTO   = "Bulto"

# Fecha: buscar en columna "Datos Varios" (col AF ~índice 31) o fallback a otras
COL_FECHA_CANDIDATES = ["Datos Varios", "Fecha Ingreso", "Fecha", "fecha", "Date"]

CB="#1E6BB5"; CB2="#4A9FE0"; CG="#10B981"; CA="#F59E0B"
CP="#8B5CF6"; CR="#EF4444"; CN="#0F1C2E"; CT="#06B6D4"

PAL = [CB, CG, CA, CP, CR, CB2, CT, "#F97316", "#84CC16", "#EC4899"]

PCFG = dict(
    template="plotly_white",
    font=dict(family="Inter,sans-serif", color=CN, size=12),
    paper_bgcolor="#FFFFFF", plot_bgcolor="#FFFFFF",
    legend=dict(orientation="h", yanchor="bottom", y=-0.28,
                xanchor="center", x=0.5, bgcolor="rgba(0,0,0,0)", font_size=11),
    hoverlabel=dict(bgcolor=CN, font_color="#FFF",
                    font_family="Inter", font_size=12, bordercolor=CN),
)
TOOLBAR = {
    "modeBarButtonsToRemove": ["select2d","lasso2d","autoScale2d"],
    "displaylogo": False,
    "toImageButtonOptions": {"format":"png","filename":"flash_chart","scale":2},
}
MES_ES = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
          7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}

def _parse_fecha_col(series: "pd.Series") -> "pd.Series":
    """Convierte una columna de fechas que puede ser string ('dd-mm-yyyy'),
    o número serial de Excel (days since 1899-12-30) a datetime."""
    import pandas as _pd
    # Intentar como string primero
    dt = _pd.to_datetime(series, dayfirst=True, errors="coerce")
    # Para celdas que no parsearon como string, intentar como serial Excel
    mask_nat = dt.isna()
    if mask_nat.any():
        nums = _pd.to_numeric(series[mask_nat], errors="coerce")
        valid_nums = nums[nums.notna() & (nums > 1)]   # seriales > 1 = fechas reales
        if not valid_nums.empty:
            # Excel serial: días desde 1899-12-30
            excel_epoch = _pd.Timestamp("1899-12-30")
            xl_dates = valid_nums.apply(
                lambda n: excel_epoch + _pd.Timedelta(days=int(n)))
            dt = dt.copy()
            dt.update(xl_dates)
    return dt

def fecha_corta(fecha_str):
    try:
        dt = pd.to_datetime(fecha_str)
        return f"{dt.day:02d}-{MES_ES[dt.month]}"
    except:
        return str(fecha_str)

# ═══════════════════════════════════════════════════════════════════════════
#  HELPERS GENERALES
# ═══════════════════════════════════════════════════════════════════════════
def safe(df, col):
    return df[col] if col in df.columns else pd.Series([""]*len(df), name=col)

def get_fecha_col(df):
    """Devuelve el nombre de la columna de fecha a usar, priorizando 'Datos Varios'."""
    for c in COL_FECHA_CANDIDATES:
        if c in df.columns:
            return c
    # Si hay columna en posición AF (índice 31)
    if len(df.columns) > 31:
        return df.columns[31]
    return None

def _read_xls_biff8_native(raw: bytes) -> "pd.DataFrame":
    """Lee un archivo XLS BIFF8/OLE2 sin xlrd.
    Funciona con los archivos que xlrd 2.x rechaza (seen[2]==4).
    Soporta SST, LABELSST, NUMBER, RK, MULRK, LABEL.
    """
    import struct as _s

    END_OF_CHAIN = 0xFFFFFFFE
    FREE_SECTOR  = 0xFFFFFFFF

    if raw[:4] != b'\xd0\xcf\x11\xe0':
        raise ValueError("No es un archivo OLE2/BIFF")

    sec_sz = 1 << _s.unpack_from('<H', raw, 30)[0]
    first_dir = _s.unpack_from('<I', raw, 48)[0]
    difat = list(_s.unpack_from('<109I', raw, 76))

    def _sector(sid):
        off = (sid + 1) * sec_sz
        return raw[off:off + sec_sz]

    fat = []
    for sid in difat:
        if sid in (FREE_SECTOR, END_OF_CHAIN, 0xFFFFFFFD): break
        fat.extend(_s.unpack_from(f'<{sec_sz//4}I', _sector(sid)))

    def _chain(start):
        ch, s = [], start
        while s not in (END_OF_CHAIN, FREE_SECTOR) and s < len(fat):
            ch.append(s); s = fat[s]
        return ch

    dir_data = b''.join(_sector(s) for s in _chain(first_dir))
    stream = None
    for i in range(len(dir_data) // 128):
        e = dir_data[i*128:(i+1)*128]
        nl = _s.unpack_from('<H', e, 64)[0]
        nm = e[:nl].decode('utf-16-le', errors='replace').rstrip('\x00')
        if nm.lower() in ('workbook', 'book') and e[66] == 2:
            ss, sz = _s.unpack_from('<I', e, 116)[0], _s.unpack_from('<I', e, 120)[0]
            stream = b''.join(_sector(s) for s in _chain(ss))[:sz]
            break

    if stream is None:
        raise ValueError("No se encontró el stream Workbook en el OLE2")

    # ── Parsear SST (con soporte de CONTINUE y grbit mid-string) ─────
    sst = []
    pos = sst_end = 0
    while pos < len(stream) - 4:
        rt = _s.unpack_from('<H', stream, pos)[0]
        rl = _s.unpack_from('<H', stream, pos + 2)[0]
        if rt == 0x00fc:                        # SST
            chunks = [stream[pos+4:pos+4+rl]]
            # Marcar posición en el body donde empieza cada CONTINUE
            # (ese primer byte es el grbit de la cadena que cruza el límite)
            bnd = set()                         # posiciones de bytes grbit en el body
            p2 = pos + 4 + rl
            while p2 < len(stream) - 4:        # absorber CONTINUE (0x003c, NO 0x00fe)
                if _s.unpack_from('<H', stream, p2)[0] != 0x003c: break
                cl = _s.unpack_from('<H', stream, p2+2)[0]
                bnd.add(sum(len(c) for c in chunks))   # posición del grbit en body
                chunks.append(stream[p2+4:p2+4+cl]); p2 += 4 + cl
            sst_end = p2
            body = b''.join(chunks)
            unique = _s.unpack_from('<I', body, 4)[0]
            bp = 8
            for _ in range(unique):
                if bp + 3 > len(body): break
                # Consumir cualquier límite CONTINUE que quede ANTES del inicio
                # de esta cadena (ocurre cuando el salto de rich-text de la
                # cadena anterior terminó exactamente en un límite).
                while bnd and min(bnd) < bp:
                    bnd.discard(min(bnd))
                slen = _s.unpack_from('<H', body, bp)[0]
                fl   = body[bp+2]; bp += 3
                is_uni = bool(fl & 1)
                nr = 0; az = 0
                if fl & 8 and bp+2 <= len(body): nr = _s.unpack_from('<H', body, bp)[0]; bp += 2
                if fl & 4 and bp+4 <= len(body): az = _s.unpack_from('<I', body, bp)[0]; bp += 4
                # Leer caracteres de a uno; en límites CONTINUE leer grbit y actualizar encoding
                buf = []
                for _ in range(slen):
                    if bp in bnd:               # límite de CONTINUE: byte grbit
                        bnd.discard(bp)
                        is_uni = bool(body[bp] & 1); bp += 1
                    if is_uni:
                        if bp + 2 > len(body): break
                        buf.append(chr(_s.unpack_from('<H', body, bp)[0])); bp += 2
                    else:
                        if bp >= len(body): break
                        buf.append(chr(body[bp])); bp += 1
                bp += nr*4 + az
                sst.append(''.join(buf))
            break
        pos += 4 + rl

    # ── Parsear celdas (después del EOF Globals) ──────────────────────────
    # max_r_dim: cota proveniente del registro DIMENSIONS.
    # Garantiza que no se pierda la última fila aunque sus celdas sean
    # exclusivamente BLANK (formateadas pero vacías) o FORMULA sin caché.
    cells = {}; max_r = max_c = 0; max_r_dim = 0

    def _put(r, c, v):
        nonlocal max_r, max_c
        cells[(r, c)] = v; max_r = max(max_r, r); max_c = max(max_c, c)

    def _row(r, c):
        """Registra fila/col sin escribir valor (celdas BLANK/FORMULA vacías)."""
        nonlocal max_r, max_c
        max_r = max(max_r, r); max_c = max(max_c, c)

    pos = sst_end if sst_end else 0
    # saltar hasta EOF del BOF Global
    while pos + 4 <= len(stream):
        rt = _s.unpack_from('<H', stream, pos)[0]
        rl = _s.unpack_from('<H', stream, pos + 2)[0]
        pos += 4 + rl
        if rt == 0x000a: break

    while pos + 4 <= len(stream):
        rt = _s.unpack_from('<H', stream, pos)[0]
        rl = _s.unpack_from('<H', stream, pos + 2)[0]
        bd = stream[pos+4:pos+4+rl]; pos += 4 + rl
        if rt == 0x000a: break

        if rt == 0x0200 and rl >= 14:                        # DIMENSIONS
            # rowmax es exclusivo (primer índice vacío tras los datos)
            max_r_dim = max(0, _s.unpack_from('<I', bd, 4)[0] - 1)

        elif rt == 0x00fd and rl >= 10:                      # LABELSST
            r, c = _s.unpack_from('<H', bd, 0)[0], _s.unpack_from('<H', bd, 2)[0]
            idx  = _s.unpack_from('<I', bd, 6)[0]
            _put(r, c, sst[idx] if idx < len(sst) else '')

        elif rt == 0x0203 and rl >= 14:                      # NUMBER
            r, c = _s.unpack_from('<H', bd, 0)[0], _s.unpack_from('<H', bd, 2)[0]
            _put(r, c, _s.unpack_from('<d', bd, 6)[0])

        elif rt == 0x027e and rl >= 10:                      # RK
            r, c = _s.unpack_from('<H', bd, 0)[0], _s.unpack_from('<H', bd, 2)[0]
            rk = _s.unpack_from('<I', bd, 6)[0]
            if rk & 2: v = float(rk >> 2)
            else:
                v = _s.unpack_from('<d', _s.pack('<Q', int(rk & ~3) << 32))[0]
            _put(r, c, v / 100.0 if rk & 1 else v)

        elif rt == 0x00bd:                                   # MULRK
            r, fc = _s.unpack_from('<H', bd, 0)[0], _s.unpack_from('<H', bd, 2)[0]
            for i in range((rl - 4) // 6):
                rk = _s.unpack_from('<I', bd, 4 + i*6 + 2)[0]
                if rk & 2: v = float(rk >> 2)
                else:
                    v = _s.unpack_from('<d', _s.pack('<Q', int(rk & ~3) << 32))[0]
                _put(r, fc + i, v / 100.0 if rk & 1 else v)

        elif rt == 0x0204 and rl >= 8:                       # LABEL (old BIFF4)
            r, c = _s.unpack_from('<H', bd, 0)[0], _s.unpack_from('<H', bd, 2)[0]
            sl = _s.unpack_from('<H', bd, 6)[0]
            _put(r, c, bd[8:8+sl].decode('latin-1', errors='replace'))

        elif rt == 0x0006 and rl >= 14:                      # FORMULA
            r, c = _s.unpack_from('<H', bd, 0)[0], _s.unpack_from('<H', bd, 2)[0]
            b6, b7 = (bd[12] if len(bd) > 12 else 0), (bd[13] if len(bd) > 13 else 0)
            if b6 == 0xff and b7 == 0xff:
                _row(r, c)                                   # texto/bool: registrar fila
            else:
                _put(r, c, _s.unpack_from('<d', bd, 6)[0])  # número cacheado

        elif rt == 0x0201 and rl >= 6:                       # BLANK
            r, c = _s.unpack_from('<H', bd, 0)[0], _s.unpack_from('<H', bd, 2)[0]
            _row(r, c)

        elif rt == 0x027f and rl >= 6:                       # MULBLANK
            r  = _s.unpack_from('<H', bd, 0)[0]
            fc = _s.unpack_from('<H', bd, 2)[0]
            lc = _s.unpack_from('<H', bd, rl-2)[0]
            for c in range(fc, lc + 1): _row(r, c)

    # Aplicar cota inferior de DIMENSIONS para no perder filas con sólo BLANKs.
    if max_r_dim > max_r:
        max_r = max_r_dim

    # Eliminar filas vacías finales: si la última fila no tiene ninguna celda
    # con dato real (resultado de BLANK/max_r_dim), la descartamos para que el
    # DataFrame no incluya una fila extra de vacíos.
    while max_r > 0 and not any(cells.get((max_r, c), '') != ''
                                 for c in range(max_c + 1)):
        max_r -= 1

    if not cells and max_r == 0:
        raise ValueError("No se encontraron celdas en el worksheet BIFF8")

    grid = [[cells.get((r, c), '') for c in range(max_c + 1)]
            for r in range(max_r + 1)]
    return pd.DataFrame(grid[1:], columns=[str(v) for v in grid[0]])


def load_file(up):
    """Lee CSV, XLSX o XLS con cadena de fallbacks robusta.

    Para .xls el orden garantiza compatibilidad con archivos que xlrd 2.x
    rechaza con 'seen[2]==4' (archivos BIFF8 con ciertos registros de formato):
      1. xlrd con formatting_info=False  → evita el bug seen[2]==4
      2. Parser BIFF8 nativo             → no depende de xlrd
      3. pd.read_html                    → XLS que en realidad es HTML
      4. pd.read_csv(sep=\t)             → XLS que en realidad es TSV
    """
    ext = up.name.split('.')[-1].lower()

    if ext == 'xlsx':
        up.seek(0)
        raw_xl = up.read()
        # Primario: openpyxl read_only=False → max_row calculado desde celdas
        # reales, no desde el atributo <dimension>. Esto evita que archivos
        # con dimension desactualizada pierdan la última fila de datos.
        try:
            from openpyxl import load_workbook as _lwb
            _wb_op = _lwb(io.BytesIO(raw_xl), read_only=False, data_only=True)
            _ws_op = _wb_op.active
            _rows  = list(_ws_op.iter_rows(
                min_row=1, max_row=_ws_op.max_row,
                max_col=_ws_op.max_column, values_only=True))
            _wb_op.close()
            if _rows:
                _hdrs = [str(h) if h is not None else f"Col{i}"
                         for i, h in enumerate(_rows[0])]
                _data = [list(r) for r in _rows[1:]]
                df_xl = pd.DataFrame(_data, columns=_hdrs)
                # Descartar solo filas donde TODOS los valores son None/NaN
                df_xl = df_xl.dropna(how='all').reset_index(drop=True)
                return df_xl
        except Exception:
            pass
        # Fallback: pandas + openpyxl
        return pd.read_excel(io.BytesIO(raw_xl), engine='openpyxl')

    elif ext == 'xls':
        up.seek(0)
        raw = up.read()

        # ── Intento 1: xlrd con formatting_info=False (evita seen[2]==4) ──
        try:
            import xlrd as _xlrd
            book = _xlrd.open_workbook(file_contents=raw,
                                       formatting_info=False)
            sh   = book.sheet_by_index(0)
            if sh.nrows < 1:
                raise ValueError("Hoja vacía")
            headers = [str(sh.cell_value(0, c)) for c in range(sh.ncols)]
            rows    = [[sh.cell_value(r, c) for c in range(sh.ncols)]
                       for r in range(1, sh.nrows)]
            return pd.DataFrame(rows, columns=headers)
        except ImportError:
            pass
        except Exception:
            pass

        # ── Intento 2: parser BIFF8 nativo (sin xlrd) ─────────────────────
        try:
            return _read_xls_biff8_native(raw)
        except Exception:
            pass

        # ── Intento 3: HTML disfrazado de .xls ────────────────────────────
        try:
            import io as _io
            dfs = pd.read_html(_io.BytesIO(raw))
            if dfs and len(dfs[0]) > 0:
                return dfs[0]
        except Exception:
            pass

        # ── Intento 4: TSV/CSV con extensión .xls ─────────────────────────
        try:
            import io as _io
            return pd.read_csv(_io.BytesIO(raw), sep='\t',
                               encoding='latin-1', low_memory=False)
        except Exception:
            pass

        raise ValueError(
            f"No se pudo leer '{up.name}'. Todos los métodos fallaron.\n"
            f"Tip: abrí el archivo en Excel y guardalo como .xlsx"
        )

    elif ext == 'csv':
        for enc in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
            try:
                up.seek(0)
                return pd.read_csv(up, low_memory=False, encoding=enc)
            except Exception:
                pass
        up.seek(0)
        return pd.read_csv(up, low_memory=False,
                           encoding='latin-1', errors='replace')

    # Extensión desconocida
    up.seek(0)
    try:
        return pd.read_excel(up, engine='openpyxl')
    except Exception:
        pass
    raise ValueError(
        f"Formato no soportado: '{up.name}'. Usá CSV, XLSX o XLS."
    )

def is_ent(s: pd.Series) -> pd.Series:
    s = s.str.lower().fillna("")
    return s.str.contains("entregado") & ~s.str.contains("no entregado")

def fmt(n):
    n = int(n)
    if n >= 1_000_000: return f"{n/1_000_000:.1f}M"
    if n >= 1_000:     return f"{n/1_000:.1f}K"
    return str(n)

def validate_df(df: pd.DataFrame) -> dict:
    required  = [COL_PIEZA, COL_ESTADO]
    total     = len(df)
    missing   = [c for c in required if c not in df.columns]
    null_rows = int(df[COL_PIEZA].isna().sum()) if COL_PIEZA in df.columns else 0
    ok_rows   = total - null_rows
    fecha_col = get_fecha_col(df)
    return {"total":total,"ok":ok_rows,"errors":null_rows,
            "missing_cols":missing,"valid":len(missing)==0 and ok_rows>0,
            "fecha_col": fecha_col}

# ── UI helpers ────────────────────────────────────────────────────────────
def kcard(label, value, sub="", color="", sub_cls="", small_val=False):
    v_cls = "kv sm" if small_val else "kv"
    st.markdown(
        f'<div class="kc {color}"><div class="kl">{label}</div>'
        f'<div class="{v_cls}">{value}</div>'
        f'{"<div class=ks "+sub_cls+">"+sub+"</div>" if sub else ""}</div>',
        unsafe_allow_html=True)

def co(title, sub=""):
    st.markdown(
        f'<div class="cw"><div class="ct">{title}</div>'
        f'{"<div class=cs>"+sub+"</div>" if sub else ""}',
        unsafe_allow_html=True)

def cc(): st.markdown("</div>", unsafe_allow_html=True)

def theme(fig, h=340, ml=10, mr=10):
    fig.update_layout(height=h, margin=dict(l=ml,r=mr,t=10,b=10), **PCFG)
    fig.update_xaxes(showgrid=True, gridcolor="#F1F5F9", linecolor="#E2E8F0", zeroline=False)
    fig.update_yaxes(showgrid=True, gridcolor="#F1F5F9", linecolor="#E2E8F0", zeroline=False)
    return fig

def pc(fig): st.plotly_chart(fig, use_container_width=True, config=TOOLBAR)

def plain_table(g: pd.DataFrame, num_cols: list, pct_cols: list = []):
    cfg = {}
    for c in num_cols: cfg[c] = st.column_config.NumberColumn(c, format="%d")
    for c in pct_cols: cfg[c] = st.column_config.NumberColumn(c, format="%.0f%%")
    st.dataframe(g, use_container_width=True, hide_index=True, column_config=cfg)

# ═══════════════════════════════════════════════════════════════════════════
#  AGREGACIONES
# ═══════════════════════════════════════════════════════════════════════════
def _agg_clientes(df):
    _df = df[safe(df,COL_CLIENTE).str.strip().ne('')].copy() if COL_CLIENTE in df.columns else df
    g = (_df.assign(_e=is_ent(safe(_df,COL_ESTADO)).astype(int),
                    _b=pd.to_numeric(safe(_df,COL_BULTO),errors="coerce").fillna(0))
            .groupby(COL_CLIENTE, as_index=False)
           .agg(Total=(COL_PIEZA,"count"), Entregados=("_e","sum"), Bultos=("_b","sum")))
    g["Pendientes"]    = g["Total"] - g["Entregados"]
    g["% Efectividad"] = (g["Entregados"]/g["Total"]*100).round(1)
    return g.sort_values("Total",ascending=False).reset_index(drop=True).rename(columns={COL_CLIENTE:"Cliente"})

def _agg_distrib(df):
    g = (df.assign(_e=is_ent(safe(df,COL_ESTADO)).astype(int),
                   _b=pd.to_numeric(safe(df,COL_BULTO),errors="coerce").fillna(0))
           .groupby(COL_DISTRIB, as_index=False)
           .agg(Total=(COL_PIEZA,"count"), Entregados=("_e","sum"), Bultos=("_b","sum"),
                HDRs=(COL_HDR,"nunique"), Clientes=(COL_CLIENTE,"nunique")))
    g["Pendientes"]    = g["Total"] - g["Entregados"]
    g["% Efectividad"] = (g["Entregados"]/g["Total"]*100).round(1)
    return g.sort_values("Total",ascending=False).reset_index(drop=True).rename(columns={COL_DISTRIB:"Distribuidor"})

def _agg_fecha(df):
    """Usa la columna de fecha correcta (Datos Varios / col AF), con fallback a Fecha Ingreso."""
    d    = df.copy()
    fcol = get_fecha_col(d)
    if fcol is None:
        d["_f"] = "Sin fecha"
    else:
        d["_dt"] = _parse_fecha_col(d[fcol])
        # Fallback a "Fecha Ingreso" (col A) para filas donde la col principal está vacía
        if "Fecha Ingreso" in d.columns and d["_dt"].isna().any():
            _fb = _parse_fecha_col(d["Fecha Ingreso"])
            d["_dt"] = d["_dt"].fillna(_fb)
        d["_f"]  = d["_dt"].dt.strftime("%Y-%m-%d").fillna("Sin fecha")
    d["_e"] = is_ent(safe(d,COL_ESTADO)).astype(int)
    d["_b"] = pd.to_numeric(safe(d,COL_BULTO), errors="coerce").fillna(0)
    g = (d.groupby("_f", as_index=False)
          .agg(Total=(COL_PIEZA,"count"), Entregados=("_e","sum"), Bultos=("_b","sum")))
    g["Pendientes"] = g["Total"] - g["Entregados"]
    return g.sort_values("_f").reset_index(drop=True).rename(columns={"_f":"Fecha"})

def _agg_dow(df):
    """Pedidos por día de la semana usando la columna correcta."""
    d    = df.copy()
    fcol = get_fecha_col(d)
    if fcol is None:
        return pd.DataFrame({"Dia":["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"],
                              "Total":[0]*7})
    d["_dt"] = _parse_fecha_col(d[fcol])
    if "Fecha Ingreso" in d.columns and d["_dt"].isna().any():
        _fb = _parse_fecha_col(d["Fecha Ingreso"])
        d["_dt"] = d["_dt"].fillna(_fb)
    d["_dow"] = d["_dt"].dt.dayofweek
    d = d.dropna(subset=["_dow"])
    dow_map = {0:"Lunes",1:"Martes",2:"Miércoles",3:"Jueves",4:"Viernes",5:"Sábado",6:"Domingo"}
    grp = d.groupby("_dow").agg(Total=(COL_PIEZA,"count")).reset_index()
    all_days = pd.DataFrame({"_dow":range(7),"Dia":[dow_map[i] for i in range(7)]})
    grp = all_days.merge(grp[["_dow","Total"]], on="_dow", how="left").fillna(0)
    grp["Total"] = grp["Total"].astype(int)
    return grp[["Dia","Total"]]

def _agg_dow_by_client(df):
    """Pedidos por día de la semana desglosado por cliente (top 8)."""
    d    = df.copy()
    fcol = get_fecha_col(d)
    dow_map = {0:"Lunes",1:"Martes",2:"Miércoles",3:"Jueves",4:"Viernes",5:"Sábado",6:"Domingo"}
    all_days = pd.DataFrame({"_dow": range(7),
                              "Dia":  [dow_map[i] for i in range(7)]})
    if fcol is None or COL_CLIENTE not in d.columns:
        return pd.DataFrame(), []
    d["_dt"] = _parse_fecha_col(d[fcol])
    if "Fecha Ingreso" in d.columns and d["_dt"].isna().any():
        _fb = _parse_fecha_col(d["Fecha Ingreso"])
        d["_dt"] = d["_dt"].fillna(_fb)
    d["_dow"] = d["_dt"].dt.dayofweek
    d = d.dropna(subset=["_dow"])
    d["_dow"] = d["_dow"].astype(int)
    # Limitar a top 8 clientes por volumen total
    top8 = (d.groupby(COL_CLIENTE)[COL_PIEZA].count()
              .sort_values(ascending=False).head(8).index.tolist())
    result = {}
    for cli in top8:
        g = (d[d[COL_CLIENTE] == cli]
               .groupby("_dow").agg(_t=(COL_PIEZA,"count")).reset_index())
        g = all_days.merge(g[["_dow","_t"]], on="_dow", how="left").fillna(0)
        g["_t"] = g["_t"].astype(int)
        result[cli] = g["_t"].tolist()   # lista de 7 valores (Lun→Dom)
    dias = all_days["Dia"].tolist()
    return result, dias   # {cliente: [7 ints]}, [7 nombres días]

def _agg_estado(df):
    d = df.copy()
    d["_b"] = pd.to_numeric(safe(d,COL_BULTO), errors="coerce").fillna(0)
    g = (d.groupby(COL_ESTADO, as_index=False)
          .agg(Total=(COL_PIEZA,"count"), Bultos=("_b","sum"),
               Clientes=(COL_CLIENTE,"nunique"), Distribuidores=(COL_DISTRIB,"nunique")))
    return g.sort_values("Total",ascending=False).reset_index(drop=True).rename(columns={COL_ESTADO:"Estado"})

def _agg_zonas(df):
    d = df.copy()
    d["_z"] = (safe(d,COL_LOC).replace("",pd.NA).fillna(safe(d,"Localidad Origen")).fillna("Sin zona"))
    d["_e"] = is_ent(safe(d,COL_ESTADO)).astype(int)
    d["_b"] = pd.to_numeric(safe(d,COL_BULTO), errors="coerce").fillna(0)
    g = (d.groupby("_z", as_index=False)
          .agg(Total=(COL_PIEZA,"count"), Entregados=("_e","sum"), Bultos=("_b","sum"),
               Clientes=(COL_CLIENTE,"nunique"), Distribuidores=(COL_DISTRIB,"nunique")))
    g["Pendientes"]    = g["Total"] - g["Entregados"]
    g["% Efectividad"] = (g["Entregados"]/g["Total"]*100).round(1)
    return g.sort_values("Total",ascending=False).reset_index(drop=True).rename(columns={"_z":"Zona"})

def _agg_tarifa(df, tarifa_map: dict):
    if not tarifa_map:
        return pd.DataFrame(columns=["Cliente","Entregados","Total a Cobrar"])
    d = df.copy()
    d["_e"]   = is_ent(safe(d, COL_ESTADO)).astype(int)
    d["_cli"] = safe(d, COL_CLIENTE).astype(str).str.strip()
    d["_loc"] = safe(d, COL_LOC).astype(str).str.strip()
    
    def get_tar(row):
        cli = row["_cli"]
        loc = row["_loc"]
        if cli in tarifa_map:
            cdata = tarifa_map[cli]
            if isinstance(cdata, dict):
                return cdata.get(loc, cdata.get("DEFAULT", 0))
            return cdata
        return 0

    d["_tar_u"]  = d.apply(get_tar, axis=1)
    d["_cobro"]  = d["_e"] * d["_tar_u"]
    g = d[d["_cli"].isin(tarifa_map.keys())].groupby("_cli", as_index=False).agg(
        Entregados=("_e", "sum"),
        Total_Cobrar=("_cobro", "sum")
    ).rename(columns={"_cli":"Cliente", "Total_Cobrar":"Total a Cobrar"})
    return g.sort_values("Total a Cobrar", ascending=False).reset_index(drop=True)


# ═══════════════════════════════════════════════════════════════════════════
#  MOTOR EXCEL — v7.0
#  Layout por hoja:
#    fila 0     → vacía (separador del borde)
#    fila 1     → título
#    fila 2     → encabezados tabla
#    filas 3..  → datos
#    fila n+5   → gráfico 1 (centrado)   col B
#    fila n+5   → gráfico 2 (centrado)   col L
#  Zoom 90%, sin freeze panes
# ═══════════════════════════════════════════════════════════════════════════
def _fmts(wb):
    F, S = "Calibri", 11
    return {
        # Títulos en azul
        "T":   wb.add_format({"bold":True,"font_name":F,"font_size":13,
                               "font_color":CB,"bg_color":"#FFFFFF","border":0}),
        # Encabezados: fondo azul, letra BLANCA
        "H":   wb.add_format({"bold":True,"font_name":F,"font_size":S,
                               "bg_color":CB,"font_color":"#FFFFFF",
                               "border":1,"border_color":"#0F4C8A",
                               "align":"center","valign":"vcenter"}),
        # Fila blanca
        "R":   wb.add_format({"font_name":F,"font_size":S,
                               "bg_color":"#FFFFFF","border":1,"border_color":"#D0DCE8","valign":"vcenter"}),
        # Fila azul claro (alternas)
        "A":   wb.add_format({"font_name":F,"font_size":S,
                               "bg_color":"#F0F7FF","border":1,"border_color":"#D0DCE8","valign":"vcenter"}),
        # Número con comas — blanco
        "N":   wb.add_format({"font_name":F,"font_size":S,"num_format":"#,##0",
                               "border":1,"border_color":"#D0DCE8","align":"right","valign":"vcenter"}),
        # Número con comas — azul
        "NA":  wb.add_format({"font_name":F,"font_size":S,"num_format":"#,##0",
                               "bg_color":"#F0F7FF","border":1,"border_color":"#D0DCE8","align":"right","valign":"vcenter"}),
        # Porcentaje — blanco
        "P":   wb.add_format({"font_name":F,"font_size":S,"num_format":'0.0"%"',
                               "border":1,"border_color":"#D0DCE8","align":"right","valign":"vcenter"}),
        # Porcentaje — azul
        "PA":  wb.add_format({"font_name":F,"font_size":S,"num_format":'0.0"%"',
                               "bg_color":"#F0F7FF","border":1,"border_color":"#D0DCE8","align":"right","valign":"vcenter"}),
        "BG":  wb.add_format({"bg_color":"#FFFFFF","border":0}),
        # Datos generales — subtítulos
        "DT":  wb.add_format({"bold":True,"font_name":F,"font_size":S,
                               "font_color":CB,"bg_color":"#FFFFFF","border":0}),
        # Encabezado tabla de datos
        "DH":  wb.add_format({"bold":True,"font_name":F,"font_size":S,
                               "bg_color":CB,"font_color":"#FFFFFF",
                               "border":1,"border_color":"#0F4C8A",
                               "align":"center","valign":"vcenter","text_wrap":True}),
        # Encabezado tabla fila 5 (wrap + vcenter)
        "DH5": wb.add_format({"bold":True,"font_name":F,"font_size":S,
                               "bg_color":CB,"font_color":"#FFFFFF",
                               "border":1,"border_color":"#0F4C8A",
                               "align":"center","valign":"vcenter","text_wrap":True}),
        # Celda blanca
        "DR":  wb.add_format({"font_name":F,"font_size":S,
                               "bg_color":"#FFFFFF","border":1,"border_color":"#D0DCE8",
                               "valign":"vcenter"}),
        # Celda azul claro
        "DA":  wb.add_format({"font_name":F,"font_size":S,
                               "bg_color":"#F0F7FF","border":1,"border_color":"#D0DCE8",
                               "valign":"vcenter"}),
        # Número entero sin comas — blanco (PIEZA,HDR,CP,TEL)
        "DNI": wb.add_format({"font_name":F,"font_size":S,"num_format":"0",
                               "bg_color":"#FFFFFF","border":1,"border_color":"#D0DCE8",
                               "align":"right","valign":"vcenter"}),
        # Número entero sin comas — azul
        "DNIA":wb.add_format({"font_name":F,"font_size":S,"num_format":"0",
                               "bg_color":"#F0F7FF","border":1,"border_color":"#D0DCE8",
                               "align":"right","valign":"vcenter"}),
        # Número con comas — blanco (Nro, otros)
        "DN":  wb.add_format({"font_name":F,"font_size":S,"num_format":"#,##0",
                               "bg_color":"#FFFFFF","border":1,"border_color":"#D0DCE8",
                               "align":"right","valign":"vcenter"}),
        # Número con comas — azul
        "DNA": wb.add_format({"font_name":F,"font_size":S,"num_format":"#,##0",
                               "bg_color":"#F0F7FF","border":1,"border_color":"#D0DCE8",
                               "align":"right","valign":"vcenter"}),
        # Porcentaje sin decimales — blanco
        "DPCT": wb.add_format({"font_name":F,"font_size":S,"num_format":"0%",
                                "bg_color":"#FFFFFF","border":1,"border_color":"#D0DCE8",
                                "align":"center","valign":"vcenter"}),
        # Porcentaje sin decimales — azul
        "DPCTa":wb.add_format({"font_name":F,"font_size":S,"num_format":"0%",
                                "bg_color":"#F0F7FF","border":1,"border_color":"#D0DCE8",
                                "align":"center","valign":"vcenter"}),
        # Valor centrado — blanco
        "DVC": wb.add_format({"font_name":F,"font_size":S,"num_format":"#,##0",
                               "bg_color":"#FFFFFF","border":1,"border_color":"#D0DCE8",
                               "align":"center","valign":"vcenter"}),
        # Valor centrado — azul
        "DVCa":wb.add_format({"font_name":F,"font_size":S,"num_format":"#,##0",
                               "bg_color":"#F0F7FF","border":1,"border_color":"#D0DCE8",
                               "align":"center","valign":"vcenter"}),
        "DP":  wb.add_format({"font_name":F,"font_size":S,"num_format":'0.0"%"',
                               "border":1,"border_color":"#D0DCE8","align":"right","valign":"vcenter"}),
        # Encabezado BG blanco (celda vacía con borde)
        "HBG": wb.add_format({"bg_color":"#FFFFFF","border":0}),
        # Formatos para tablas Datos Generales (sin bg_color para banding)
        "T_FEC": wb.add_format({"font_name":F,"font_size":S,"num_format":"dd/mm/yy","valign":"vcenter","align":"center"}),
        "T_MON": wb.add_format({"font_name":F,"font_size":S,"num_format":"$ #,##0","valign":"vcenter","align":"right"}),
        "T_INT": wb.add_format({"font_name":F,"font_size":S,"num_format":"0","valign":"vcenter","align":"right"}),
    }

def _setup_ws(ws, fmts, zoom=90):
    """Configura hoja: fondo blanco, zoom 90%, sin freeze."""
    ws.set_zoom(zoom)
    for row_i in range(300):
        ws.set_row(row_i, None, fmts["BG"])
    ws.set_column(0, 40, None, fmts["BG"])

def _write_module_sheet(wb, fmts, sheet_name, df, title, no_table=False):
    """
    Escribe tabla en hoja de módulo usando add_table con Table Style Medium 2.
      fila 0  → margen
      fila 1  → título (azul)
      fila 2  → encabezado de tabla (header row de add_table)
      filas 3..2+n → datos (gestionados por la tabla)
    Retorna (ws, col_map, n_rows).
    Sin bg_color en celdas de datos: el estilo de tabla aplica el banding.
    no_table=True: sólo crea la hoja y el título, sin llamar add_table
                   (usar cuando la hoja llamante gestiona su propio add_table).
    """
    ws = wb.add_worksheet(sheet_name)
    _setup_ws(ws, fmts)
    ws.set_row(0, 8,  fmts["BG"])
    ws.set_row(1, 24, fmts["BG"])

    ncols  = len(df.columns)
    n      = len(df)
    col_map = {col: ci+1 for ci, col in enumerate(df.columns)}
    pct_cols = {c for c in df.columns if "%" in c}

    ws.merge_range(1, 1, 1, ncols, title, fmts["T"])

    if n == 0:
        # Sin datos: sólo encabezados manuales
        for ci, col in enumerate(df.columns):
            ws.write(2, ci+1, col, fmts["H"])
        ws.set_column(0, 0, 2)
        return ws, col_map, 0

    # ── Formatos de celda sin bg_color (el estilo de tabla gestiona el banding)
    F, S = "Calibri", 11
    _fmt_s  = wb.add_format({"font_name":F,"font_size":S,"valign":"vcenter"})
    _fmt_n  = wb.add_format({"font_name":F,"font_size":S,"num_format":"#,##0",
                              "align":"right","valign":"vcenter"})
    _fmt_p  = wb.add_format({"font_name":F,"font_size":S,"num_format":'0"%"',
                              "align":"right","valign":"vcenter"})

    # ── Pre-procesar datos (float entero → int)
    tbl_data = []
    for ri in range(n):
        row_vals = []
        for col in df.columns:
            v = df.iloc[ri][col]
            if not isinstance(v, str) and pd.isna(v):
                row_vals.append("")
            elif isinstance(v, float) and v == int(v) and col not in pct_cols:
                row_vals.append(int(v))
            else:
                row_vals.append(v)
        tbl_data.append(row_vals)

    # ── Definición de columnas
    tbl_cols = []
    for col in df.columns:
        if col in pct_cols:
            fmt = _fmt_p
        elif df[col].dtype.kind in ("i","f"):
            fmt = _fmt_n
        else:
            fmt = _fmt_s
        tbl_cols.append({
            "header"        : col,
            "header_format" : fmts["H"],
            "format"        : fmt,
        })

    tbl_name = "Tabla" + "".join(c for c in sheet_name if c.isalnum())
    if not no_table:
        ws.add_table(2, 1, 2 + n, ncols, {
            "name"        : tbl_name,
            "style"       : "Table Style Medium 2",
            "autofilter"  : False,
            "banded_rows" : True,
            "first_column": False,
            "columns"     : tbl_cols,
            "data"        : tbl_data,
        })

    # ── Ancho de columnas
    ws.set_column(0, 0, 2)
    for ci, col in enumerate(df.columns):
        w = max(len(col), df[col].astype(str).str.len().max() if n > 0 else 0) + 3
        ws.set_column(ci+1, ci+1, min(w, 36))

    return ws, col_map, n

def _cs(wb, chart, legend="bottom"):
    chart.set_style(10)
    chart.set_chartarea({"border":{"none":True},"fill":{"color":"#FFFFFF"}})
    chart.set_plotarea({"border":{"color":"#E2E8F0","width":0.75},"fill":{"color":"#FFFFFF"}})
    chart.set_legend({"position":legend,"font":{"name":"Arial","size":10},"border":{"none":True}})
    return chart

def _tf():  return {"name_font":{"name":"Arial","size":12,"bold":True}}
def _af(lbl=""): return {"name":lbl,"name_font":{"name":"Arial","size":9},
                         "major_gridlines":{"visible":True,"line":{"color":"#E8EDF2"}}}

def _dl(val_with_labels=True):
    """Data labels config para series de columnas/barras."""
    if val_with_labels:
        return {"value":True,"font":{"name":"Arial","size":8,"bold":True}}
    return None

def build_full_excel(df: pd.DataFrame, tarifa_map: dict = None) -> bytes:
    if tarifa_map is None: tarifa_map = {}
    buf = io.BytesIO()
    wb  = xlsxwriter.Workbook(buf, {"in_memory":True})
    fmts = _fmts(wb)

    # ══════════════════════════════════════════════════════════════
    # HOJA 0: DATOS GENERALES
    # ══════════════════════════════════════════════════════════════
    ws_d = wb.add_worksheet("Datos Generales")
    _setup_ws(ws_d, fmts)
    ws_d.set_zoom(90)
    # Filas 0-3: fondo blanco sin cuadrículas
    for rr, h in enumerate([8, 28, 10, 20]):
        ws_d.set_row(rr, h, fmts["BG"])
    # Fila 4 (encabezados tabla de datos): alto=45, SOLO esta fila con wrap+center
    hdr5_fmt = wb.add_format({"bold":True,"font_name":"Calibri","font_size":11,
                               "bg_color":CB,"font_color":"#FFFFFF",
                               "border":1,"border_color":"#0F4C8A",
                               "align":"center","valign":"vcenter","text_wrap":True})
    ws_d.set_row(4, 45)

    # Col A angosta, col D separador=1.7
    ws_d.set_column(0, 0, 3.29, fmts["BG"])
    ws_d.set_column(3, 3, 1.7,  fmts["BG"])

    # Título general — SIN merge, solo write en celda B2, fuente 13
    titulo_fmt = wb.add_format({"bold":True,"font_name":"Calibri","font_size":13,
                                 "font_color":CB,"bg_color":"#FFFFFF","border":0})
    ws_d.write(1, 1, "Flash Expres - Analytics — Datos Generales", titulo_fmt)

    # ── Resumen Ejecutivo ──────────────────────────────────────────────────────
    subtit_fmt = wb.add_format({"bold":True,"font_name":"Calibri","font_size":13,
                                 "font_color":CB,"bg_color":"#FFFFFF","border":0})
    ws_d.write(3, 1, "Resumen Ejecutivo", subtit_fmt)
    ws_d.set_column(1, 1, 22)
    ws_d.set_column(2, 2, 10)

    total_ped = len(df)
    n_ent_g   = int(is_ent(safe(df,COL_ESTADO)).sum())
    pct_ent_g = round(n_ent_g/total_ped*100,1) if total_ped else 0
    n_cli_g   = df[COL_CLIENTE].nunique() if COL_CLIENTE in df.columns else 0
    n_drv_g   = df[COL_DISTRIB].nunique() if COL_DISTRIB in df.columns else 0
    n_zon_g   = safe(df,COL_LOC).replace("",pd.NA).dropna().nunique()
    n_bul_g   = int(pd.to_numeric(safe(df,COL_BULTO),errors="coerce").fillna(0).sum())

    resumen = [
        ("Total de Pedidos",       total_ped,         False),
        ("Pedidos Entregados",     n_ent_g,            False),
        ("Efectividad Global",     pct_ent_g/100.0,    True),
        ("Total Clientes",         n_cli_g,            False),
        ("Total Distribuidores",   n_drv_g,            False),
        ("Total Zonas de Entrega", n_zon_g,            False),
        ("Total Bultos",           n_bul_g,            False),
    ]

    # Formatos para celdas de datos del Resumen (se escriben DESPUÉS de add_table
    # sin data=, lo que es seguro: sólo los encabezados están en la tabla).
    res_hdr_fmt = wb.add_format({
        "bold":True,"font_name":"Calibri","font_size":11,
        "font_color":"#FFFFFF","border":1,"border_color":"#2F5496",
        "align":"center","valign":"vcenter",
    })
    res_lbl_fmt = wb.add_format({
        "font_name":"Calibri","font_size":11,"valign":"vcenter",
    })
    res_num_fmt = wb.add_format({
        "font_name":"Calibri","font_size":11,"num_format":"#,##0",
        "align":"center","valign":"vcenter",
    })
    res_pct_fmt = wb.add_format({
        "font_name":"Calibri","font_size":11,"num_format":"0%",
        "align":"center","valign":"vcenter",
    })

    # add_table sin data= → encabezados via header_format, filas escritas separado.
    # Esto permite formatos mixtos por fila (número / porcentaje) sin corromper XML.
    n_res = len(resumen)
    ws_d.add_table(4, 1, 4 + n_res, 2, {
        "name"        : "TablaResumen",
        "style"       : "Table Style Medium 2",
        "banded_rows" : True,
        "autofilter"  : False,
        "columns"     : [
            {"header": "Métrica", "header_format": res_hdr_fmt},
            {"header": "Valor",   "header_format": res_hdr_fmt},
        ],
    })
    # Escribir datos del resumen (filas 5..4+n_res, cols 1-2)
    for ri, (lbl, val, is_pct) in enumerate(resumen, start=5):
        ws_d.write(ri, 1, lbl, res_lbl_fmt)
        ws_d.write_number(ri, 2, float(val), res_pct_fmt if is_pct else res_num_fmt)

    # ── Tabla de Datos (col 4 en adelante) — formato tabla dinámica ──
    DT_HEADERS  = ["PIEZA","ID","HDR","FECHA","DRIVER","CLIENTE",
                   "DOMICILIO\nORIGEN","LOCALIDAD\nORIGEN","CP\nORIGEN",
                   "DESTINATARIO","DOMICILIO\nDESTINO","LOCALIDAD\nDESTINO",
                   "CP\nDESTINO","VALOR\nDECLARADO","TARIFA\nCLIENTE"]
    DT_HDRS_CLEAN = ["PIEZA","ID","HDR","FECHA","DRIVER","CLIENTE",
                     "DOM ORIGEN","LOC ORIGEN","CP ORIGEN",
                     "DESTINATARIO","DOM DESTINO","LOC DESTINO",
                     "CP DESTINO","VALOR DECLARADO","TARIFA CLIENTE"]
    # AT = columna 46 fuente (1-based) -> idx 45 (0-based)
    DT_SRC_IDX   = [1, 2, 9, 31, 8, 6, 17, 19, 18, 16, 21, 23, 22, 45, None]
    # hi=7 LOC ORIGEN y hi=11 LOC DESTINO -> ancho 8; hi=13,14 misma anchura TARIFA
    DT_WIDTHS    = [8.43, 12.30, 8, 9.6, 10.30, 11.30, 11, 8, 7.86,
                    14.7, 10.7, 8, 7.86, 11.57, 11.5]
    DT_INT_COLS   = {0, 2, 8, 12}
    DT_ARS_COL    = 14       # TARIFA CLIENTE
    DT_VALDEC_COL = 13       # VALOR DECLARADO

    DT_START_COL = 4
    n_dt_cols    = len(DT_HEADERS)

    # ── Formatos de columna para la tabla DG ───────────────────────────────
    F_TBL, S_TBL = "Calibri", 11

    ws_d.write(3, DT_START_COL, "Tabla de Datos", subtit_fmt)

    # Configurar anchos y formatos de columna base
    for hi, w in enumerate(DT_WIDTHS):
        col_idx = DT_START_COL + hi
        fmt = None
        if hi == 3:                   fmt = fmts["T_FEC"]
        elif hi in [13, 14]:          fmt = fmts["T_MON"]
        elif hi in DT_INT_COLS:       fmt = fmts["T_INT"]
        ws_d.set_column(col_idx, col_idx, w, fmt)

    fcol     = get_fecha_col(df)
    ncols_df = len(df.columns)

    # NOTA: encabezados de la tabla DG se aplican vía header_format en add_table.
    # No se escribe ninguna celda en la fila 4 col E..R antes de add_table.

    # ── Saneamiento: reemplazar NaN por "" para evitar corrupción XML ──
    # Esto previene el error "archivo dañado" al abrir en Excel.
    df_export = df.copy()
    df_export.fillna("", inplace=True)

    # ── add_table para Datos Generales ──
    # last_row = exactamente len(df) en índice 0-based respecto al encabezado en fila 4.
    # xlsxwriter añade la fila de totales automáticamente cuando total_row=True;
    # NO sumar +1 manualmente al last_row.
    # xlsxwriter add_table semántica con total_row=True:
    #   first_row = 4 → fila de encabezado (Excel fila 5)
    #   last_row  → fila de TOTALES (xlsxwriter la incluye dentro del rango)
    #   filas de datos = [first_row+1 .. last_row-1]
    #   → last_row = 4 + total_ped + 1  para tener total_ped filas de datos
    last_data_row = 4 + total_ped + 1  # = fila de totales; datos en filas 5..4+total_ped

    # Columnas: total_function declaradas para que add_table calcule sola
    # (Bloque de columnas eliminado por redundancia, se define abajo en table_columns_final)

    # ── Pre-computar array de datos para pasarlo a add_table ─────────────
    # Pasar data= a add_table garantiza que xlsxwriter controla todo el
    # banding, los filtros y la fila de totales sin interferencia.
    # NO se escribe ninguna celda dentro del rango de la tabla después.
    tbl_data = []
    for ri in range(total_ped):
        row_d = df_export.iloc[ri]
        row_out = []
        for hi, src_idx in enumerate(DT_SRC_IDX):
            if hi == DT_VALDEC_COL:
                # VALOR DECLARADO: leer columna AT (idx 45) como entero sin decimales
                if DT_SRC_IDX[hi] is not None and DT_SRC_IDX[hi] < ncols_df:
                    raw_v = row_d.iloc[DT_SRC_IDX[hi]]
                    try:
                        if isinstance(raw_v, (int, float)) and not pd.isna(raw_v):
                            ival = int(raw_v)
                            row_out.append(ival if ival != 0 else "")
                        else:
                            ival = int(float(str(raw_v).replace(",","").strip()))
                            row_out.append(ival if ival != 0 else "")
                    except:
                        row_out.append("")
                else:
                    row_out.append("")
            elif hi == DT_ARS_COL:
                cli_val = ""
                loc_val = ""
                if COL_CLIENTE in df_export.columns:
                    cli_val = str(df_export.iloc[ri][COL_CLIENTE]).strip()
                if COL_LOC in df_export.columns:
                    loc_val = str(df_export.iloc[ri][COL_LOC]).strip()
                
                tarifa_val = 0
                if cli_val in tarifa_map:
                    cdata = tarifa_map[cli_val]
                    if isinstance(cdata, dict):
                        tarifa_val = cdata.get(loc_val, cdata.get("DEFAULT", 0))
                    else:
                        tarifa_val = cdata
                        
                estado_val = safe(df_export, COL_ESTADO).iloc[ri]
                if tarifa_val and is_ent(pd.Series([estado_val])).any():
                    row_out.append(tarifa_val)
                else:
                    row_out.append("")
            elif hi == 3:  # FECHA
                raw_v = ""
                if fcol and fcol in df_export.columns:
                    raw_v = row_d[fcol]
                elif src_idx is not None and src_idx < ncols_df:
                    raw_v = row_d.iloc[src_idx]
                try:
                    dt_v = pd.to_datetime(raw_v, dayfirst=True, errors="coerce")
                    if not pd.isna(dt_v):
                        row_out.append(dt_v.to_pydatetime())
                    else:
                        row_out.append(str(raw_v))
                except:
                    row_out.append(str(raw_v))
            elif src_idx is not None and src_idx < ncols_df:
                raw_v = row_d.iloc[src_idx]
                if not isinstance(raw_v, str) and pd.isna(raw_v):
                    row_out.append("")
                elif hi in DT_INT_COLS:
                    try:
                        # Usar conversión directa si ya es numérico para evitar
                        # que replace(".","") sobre "20690.0" produzca "206900".
                        if isinstance(raw_v, (int, float)):
                            ival = int(raw_v)
                        else:
                            ival = int(float(str(raw_v).replace(",","").strip()))
                        # Cero indica celda vacía numérica en XLS → omitir
                        row_out.append(ival if ival != 0 else "")
                    except:
                        row_out.append(str(raw_v) if str(raw_v).strip() else "")
                elif isinstance(raw_v, (int, float)) and not isinstance(raw_v, bool):
                    row_out.append(float(raw_v))
                else:
                    row_out.append(str(raw_v))
            else:
                row_out.append("")
        tbl_data.append(row_out)

    # ── Formatos de columna para la tabla DG ─────────────────────────────────
    # TODAS las columnas necesitan un formato explícito (aunque sea mínimo).
    # Sin formato, las celdas de texto se escriben sin atributo s= en el XML,
    # y Excel no aplica el DXF banding de la tabla → fondo blanco en esas columnas.
    fmt_str_col = wb.add_format({"font_name": F_TBL, "font_size": S_TBL, "valign": "vcenter"})
    fmt_int_col = wb.add_format({"num_format": "0",        "font_name": F_TBL, "font_size": S_TBL, "valign": "vcenter"})
    fmt_id_col  = wb.add_format({"num_format": "0",        "font_name": F_TBL, "font_size": S_TBL, "align": "center", "valign": "vcenter"})
    fmt_ars_col = wb.add_format({"num_format": "$ #,##0",  "font_name": F_TBL, "font_size": S_TBL, "valign": "vcenter"})
    # CP DESTINO (hi=12): int con alineación derecha para total_string "Total a Cobrar"
    fmt_cp_dest = wb.add_format({"num_format": "0",        "font_name": F_TBL, "font_size": S_TBL, "align": "right", "valign": "vcenter"})
    fmt_date_col= wb.add_format({"num_format": "dd/mm/yy", "font_name": F_TBL, "font_size": S_TBL, "align": "center", "valign": "vcenter"})

    # header_format aplica color azul al encabezado vía XML de tabla.
    # NUNCA escribir celdas del rango de la tabla después de add_table.
    table_columns_final = []
    for hi, hdr in enumerate(DT_HDRS_CLEAN):
        col_def = {
            "header"        : hdr,
            "header_format" : hdr5_fmt,
            "format"        : fmt_str_col,   # formato base para TODAS (garantiza s= en XML)
        }
        # total_function / total_string según columna
        if hi == 0:
            col_def["total_function"] = "count"
        elif hi == 1:
            col_def["total_string"] = "Total Pedidos"
            col_def["format"] = fmt_id_col      # ID → centrado
        elif hi == 3:
            col_def["format"] = fmt_date_col    # FECHA → dd/mm/yy
        elif hi == 12:
            # CP DESTINO: fila totales vacía (ni suma ni texto)
            col_def["format"] = fmt_cp_dest     # int + align right
        elif hi == DT_VALDEC_COL:               # 13: VALOR DECLARADO → datos $ #,##0; total = texto derecha
            col_def["total_string"] = "Total a Cobrar"
            col_def["format"] = fmt_ars_col
        elif hi == DT_ARS_COL:                  # 14: TARIFA CLIENTE → $ #,##0
            col_def["total_function"] = "sum"
            col_def["format"] = fmt_ars_col

        # Formato entero para columnas numéricas
        # (excluir CP DESTINO=12, VALOR DECLARADO=DT_VALDEC_COL, TARIFA=DT_ARS_COL)
        if hi in DT_INT_COLS and hi not in {12, DT_VALDEC_COL, DT_ARS_COL}:
            col_def["format"] = fmt_int_col

        table_columns_final.append(col_def)

    ws_d.add_table(4, DT_START_COL, last_data_row, DT_START_COL + n_dt_cols - 1, {
        "name"        : "TablaDatosGenerales",
        "style"       : "Table Style Medium 2",
        "banded_rows" : True,
        "autofilter"  : True,
        "total_row"   : True,
        "columns"     : table_columns_final,
        "data"        : tbl_data,
    })
    # Sobreescribir celda "Total Pedidos" en fila de totales con alineación izquierda.
    # xlsxwriter permite write() sobre celdas individuales dentro del rango de tabla.
    fmt_id_total_left = wb.add_format({
        "font_name": F_TBL, "font_size": S_TBL,
        "align": "left", "valign": "vcenter", "bold": True,
    })
    ws_d.write(last_data_row, DT_START_COL + 1, "Total Pedidos", fmt_id_total_left)
    # Post-write "Total a Cobrar" en columna VALOR DECLARADO alineado a la derecha
    fmt_valdec_total = wb.add_format({
        "font_name": F_TBL, "font_size": S_TBL,
        "align": "right", "valign": "vcenter", "bold": True,
    })
    ws_d.write(last_data_row, DT_START_COL + DT_VALDEC_COL, "Total a Cobrar", fmt_valdec_total)


    # ═══════════════════════════════════════════════════════════════
    # HOJAS 1-5: módulos de análisis — protegidas con try/except
    # Si una aggregation falla (columna inexistente, etc.) se escribe
    # una hoja "Aviso" con el mensaje de error en lugar de crashear.
    # ═══════════════════════════════════════════════════════════════
    try:
        # ══════════════════════════════════════════════════════════════
        # HOJA 1: CLIENTES
        # ══════════════════════════════════════════════════════════════
        # Incluir columna Total Tarifa en hoja Clientes si hay tarifas
        g_cli = _agg_clientes(df); n_c = len(g_cli)
        g_tar_xl = _agg_tarifa(df, tarifa_map)
        if not g_tar_xl.empty:
            tar_idx = g_tar_xl.set_index("Cliente")["Total a Cobrar"]
            g_cli["Total Tarifa"] = g_cli["Cliente"].map(tar_idx).fillna(0).astype(int)
        ws_c, cm_c, n_c = _write_module_sheet(wb, fmts, "Clientes", g_cli,
            f"Reporte por Cliente ({n_c} clientes)", no_table=True)
        R1_c, RN_c = 3, 2+n_c
        ci_ck  = cm_c["Cliente"]; ci_ce = cm_c.get("Entregados"); ci_ct = cm_c.get("Total",2)
        ci_ctar = cm_c.get("Total Tarifa")

        # ── Tabla Clientes con add_table(data=...) ───────────────────────
        # Sin autofilter. Banding y formatos gestionados 100% por xlsxwriter.
        # La columna Total Tarifa usa mismo formato $ que las columnas numéricas.
        n_cli_cols = len(g_cli.columns)

        # Pre-computar array de datos
        cli_data = []
        fmt_cli_str = wb.add_format({"font_name": "Calibri", "font_size": 11, "valign": "vcenter"})
        fmt_cli_num = wb.add_format({"num_format": "#,##0",   "font_name": "Calibri", "font_size": 11, "align": "right", "valign": "vcenter"})
        fmt_cli_ars = wb.add_format({"num_format": "$ #,##0", "font_name": "Calibri", "font_size": 11, "align": "right", "valign": "vcenter"})
        for ri_c in range(n_c):
            row_c = []
            for ci_c, col_name in enumerate(g_cli.columns):
                val = g_cli.iloc[ri_c, ci_c]
                if isinstance(val, float) and val == int(val):
                    val = int(val)
                row_c.append(val if not (isinstance(val, float) and pd.isna(val)) else 0)
            cli_data.append(row_c)

        # Definición de columnas con formatos
        tbl_cols_cli = []
        for col_name in g_cli.columns:
            col_def = {
                "header"        : col_name,
                "header_format" : fmts["H"],   # formato azul vía tabla XML, sin post-write
            }
            if col_name == "Total Tarifa":
                col_def["format"] = fmt_cli_ars
            elif "%" in col_name:
                col_def["format"] = wb.add_format({"font_name":"Calibri","font_size":11,
                    "num_format":'0"%"',"align":"right","valign":"vcenter"})
            elif col_name == "Cliente":
                col_def["format"] = fmt_cli_str   # formato mínimo → garantiza s= en XML → banding OK
            else:
                col_def["format"] = fmt_cli_num
            tbl_cols_cli.append(col_def)

        ws_c.add_table(2, 1, 2 + n_c, n_cli_cols, {
            "name"        : "TablaClientes",
            "style"       : "Table Style Medium 2",
            "autofilter"  : False,
            "banded_rows" : True,
            "first_column": False,
            "columns"     : tbl_cols_cli,
            "data"        : cli_data,
        })
        # header_format ya está en tbl_cols_cli — NO re-escribir encabezados.

        # ── Gráficos debajo de la tabla — posición dinámica ──
        # start_row = len(df_clientes) + 4 como piden las directrices
        # En xlsxwriter insert_chart acepta (row, col) 0-indexed.
        # Los datos empiezan en fila 3 (encabezado=2), así que:
        # start_row (0-indexed) = n_c + 4  → equivale a len(g_cli) + 4
        chart_row_c = n_c + 4   # start_row dinámica debajo de tabla

        # Tamaño exacto según directrices: width=320, height=300
        # G1 en col B (=1), G2 en col H (=7), G3 en col N (=13) → ~2 cols vacías entre c/u
        # ── Paleta idéntica a la web ──────────────────────────────────────
        _pal_cli = ["#1E6BB5","#10B981","#F59E0B","#8B5CF6","#EF4444",
                    "#4A9FE0","#06B6D4","#F97316","#84CC16","#EC4899"]
        top15_rows = min(15, n_c)
        _top8_rows = min(8,  n_c)

        # G1 — Pedidos por Cliente: barra horizontal, sin leyenda, color por barra
        # Una sola serie + points[] = colores individuales por barra (método correcto xlsxwriter)
        ch_c1 = wb.add_chart({"type": "bar"})
        ch_c1.set_style(10)
        ch_c1.set_chartarea({"border":{"none":True},"fill":{"color":"#FFFFFF"}})
        ch_c1.set_plotarea({"border":{"color":"#E2E8F0","width":0.75},"fill":{"color":"#FFFFFF"}})
        ch_c1.set_legend({"none": True})
        ch_c1.set_x_axis({"name": "", "major_gridlines": {"visible": True,
                           "line": {"color": "#E8EDF2"}}, "num_font": {"size": 9}})
        ch_c1.set_y_axis({"name": "", "line": {"none": True},
                           "num_font": {"size": 9}})
        if ci_ce:
            ch_c1.add_series({
                "categories": ["Clientes", R1_c, ci_ck, R1_c + top15_rows - 1, ci_ck],
                "values":     ["Clientes", R1_c, ci_ce, R1_c + top15_rows - 1, ci_ce],
                "border": {"none": True},
                "data_labels": {"value": True,
                                "font": {"name": "Arial", "size": 9, "bold": True}},
                "points": [{"fill": {"color": _pal_cli[i % len(_pal_cli)]}}
                           for i in range(top15_rows)],
            })
        ch_c1.set_size({"width": 520, "height": max(380, top15_rows * 44 + 100)})
        ws_c.insert_chart(chart_row_c, 1, ch_c1)

        # G2 — Participación: donut, sin leyenda, etiquetas label+%, colores PAL
        rn8_c = min(2 + 8, 2 + n_c)
        ch_c2 = wb.add_chart({"type": "doughnut"})
        ch_c2.set_style(10)
        ch_c2.set_chartarea({"border":{"none":True},"fill":{"color":"#FFFFFF"}})
        ch_c2.set_plotarea({"border":{"color":"#E2E8F0","width":0.75},"fill":{"color":"#FFFFFF"}})
        ch_c2.set_legend({"none": True})
        ch_c2.add_series({
            "name":       "Pedidos",
            "categories": ["Clientes", R1_c, ci_ck, rn8_c, ci_ck],
            "values":     ["Clientes", R1_c, ci_ct, rn8_c, ci_ct],
            "data_labels": {"percentage": True, "category": True, "separator": "\n",
                            "font": {"name": "Arial", "size": 9}},
            "points": [{"fill": {"color": _pal_cli[i]}} for i in range(_top8_rows)],
        })
        ch_c2.set_hole_size(52)
        ch_c2.set_size({"width": 480, "height": max(380, top15_rows * 44 + 100)})
        ws_c.insert_chart(chart_row_c, 9, ch_c2)

        # G3 — Total Tarifa: barra horizontal, sin leyenda, color verde CG
        if ci_ctar:
            ch_c3 = wb.add_chart({"type": "bar"})
            ch_c3.set_style(10)
            ch_c3.set_chartarea({"border":{"none":True},"fill":{"color":"#FFFFFF"}})
            ch_c3.set_plotarea({"border":{"color":"#E2E8F0","width":0.75},"fill":{"color":"#FFFFFF"}})
            ch_c3.set_legend({"none": True})
            ch_c3.set_x_axis({"name": "", "num_format": "$ #,##0",
                               "major_gridlines": {"visible": True,
                               "line": {"color": "#E8EDF2"}}, "num_font": {"size": 9}})
            ch_c3.set_y_axis({"name": "", "line": {"none": True}, "num_font": {"size": 9}})
            ch_c3.add_series({
                "name":       "Total Tarifa",
                "categories": ["Clientes", R1_c, ci_ck, RN_c, ci_ck],
                "values":     ["Clientes", R1_c, ci_ctar, RN_c, ci_ctar],
                "fill":   {"color": "#10B981"}, "border": {"none": True},
                "data_labels": {"value": True, "num_format": "$ #,##0",
                                "font": {"name": "Arial", "size": 8, "bold": True}},
                "points": [{"fill": {"color": _pal_cli[i % len(_pal_cli)]}}
                           for i in range(n_c)],
            })
            ch_c3.set_size({"width": 460, "height": max(320, n_c * 44 + 100)})
            _g3_row = chart_row_c + (max(380, top15_rows * 44 + 100) // 20) + 3
            ws_c.insert_chart(_g3_row, 1, ch_c3)

        # ══════════════════════════════════════════════════════════════
        # HOJA 2: DISTRIBUIDORES
        # ══════════════════════════════════════════════════════════════
        g_drv = _agg_distrib(df); n_d = len(g_drv)
        ws_dr, cm_d, _ = _write_module_sheet(wb, fmts, "Distribuidores", g_drv,
            f"Reporte por Distribuidor ({n_d} drivers)")
        R1_d, RN_d = 3, 2+n_d
        ci_dk = cm_d["Distribuidor"]; ci_de = cm_d.get("Entregados"); ci_dpct = cm_d.get("% Efectividad")
        chart_row_d = n_d + 5

        ch_d1 = _cs(wb, wb.add_chart({"type":"bar"}))
        ch_d1.set_title({"name":"Pedidos Entregados por Distribuidor",**_tf()})
        ch_d1.set_x_axis(_af("Pedidos"))
        ch_d1.set_y_axis({"name":"","line":{"none":True},"name_font":{"name":"Arial","size":9}})
        if ci_de:
            ch_d1.add_series({"name":"Entregados","categories":["Distribuidores",R1_d,ci_dk,RN_d,ci_dk],
                               "values":["Distribuidores",R1_d,ci_de,RN_d,ci_de],
                               "fill":{"color":CB},"border":{"none":True},
                               "data_labels":{"value":True,"font":{"name":"Arial","size":8,"bold":True}}})
        ch_d1.set_size({"width":520,"height":max(240,n_d*38+80)})
        ws_dr.insert_chart(chart_row_d, 1, ch_d1)

        if ci_dpct:
            ch_d2 = _cs(wb, wb.add_chart({"type":"column"}))
            ch_d2.set_title({"name":"% Efectividad por Distribuidor",**_tf()})
            # Eje Y: sin título, rango 0-100, formato sin decimales
            ch_d2.set_y_axis({"name":"","min":0,"max":100,
                               "num_format":'0"%"',
                               "major_gridlines":{"visible":True,"line":{"color":"#E8EDF2"}}})
            # Eje X: sin título
            ch_d2.set_x_axis({"name":"","line":{"none":True},
                               "name_font":{"name":"Arial","size":9}})
            ch_d2.add_series({"name":"% Efectividad",
                               "categories":["Distribuidores",R1_d,ci_dk,RN_d,ci_dk],
                               "values":["Distribuidores",R1_d,ci_dpct,RN_d,ci_dpct],
                               "fill":{"color":CG},"border":{"none":True},
                               "data_labels":{"value":True,"num_format":'0"%"',
                                              "font":{"name":"Arial","size":8,"bold":True}}})
            ch_d2.set_size({"width":460,"height":260})
            ws_dr.insert_chart(chart_row_d, 10, ch_d2)

        # ══════════════════════════════════════════════════════════════
        # HOJA 3: FECHAS — tabla principal vertical + tabla DOW horizontal
        # ══════════════════════════════════════════════════════════════
        g_fec = _agg_fecha(df); n_f = len(g_fec)
        g_dow = _agg_dow(df)

        ws_f = wb.add_worksheet("Fechas")
        _setup_ws(ws_f, fmts)
        # Filas 0-1: fondo blanco explícito sin cuadrículas
        ws_f.set_row(0, 8,  fmts["BG"])
        ws_f.set_row(1, 24, fmts["BG"])
        ws_f.set_row(2, 18)
        ws_f.merge_range(1, 1, 1, 6, f"Reporte por Fecha ({n_f} días)", fmts["T"])

        # ── Tabla principal (vertical) con add_table Table Style Medium 2 ──
        fec_cols = ["Fecha","Total","Entregados","Bultos","Pendientes"]
        _Ff, _Fs = "Calibri", 11
        _fec_str = wb.add_format({"font_name":_Ff,"font_size":_Fs,"valign":"vcenter"})
        _fec_num = wb.add_format({"font_name":_Ff,"font_size":_Fs,"num_format":"#,##0",
                                   "align":"right","valign":"vcenter"})
        fec_data = []
        for _ri in range(n_f):
            _row = g_fec.iloc[_ri]
            _rowv = []
            for _col in fec_cols:
                _v = _row[_col] if _col in _row.index else ""
                if not isinstance(_v, str) and pd.isna(_v):
                    _rowv.append("")
                elif isinstance(_v, float) and _v == int(_v):
                    _rowv.append(int(_v))
                else:
                    _rowv.append(_v)
            fec_data.append(_rowv)
        fec_tbl_cols = [{"header":c,"header_format":fmts["H"],
                          "format":_fec_str if c=="Fecha" else _fec_num}
                         for c in fec_cols]
        ws_f.add_table(2, 1, 2+n_f, len(fec_cols), {
            "name":"TablaFechas","style":"Table Style Medium 2",
            "autofilter":False,"banded_rows":True,"first_column":False,
            "columns":fec_tbl_cols,"data":fec_data,
        })
        ws_f.set_column(0,0,2); ws_f.set_column(1,1,12)
        for i in range(2,6): ws_f.set_column(i,i,11)

        # ── Tabla DOW por Cliente (reemplaza la tabla DOW horizontal) ──────
        # Columnas: Cliente | Lunes | Martes | Miércoles | Jueves | Viernes | Sábado | Domingo
        # Una fila por cliente, vinculada al gráfico multi-línea.
        DOW_START_COL = 7   # col H
        _days_es = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]
        dow_by_cli, _dias_es = _agg_dow_by_client(df)
        cli_list_dow = list(dow_by_cli.keys())
        n_cli_dow    = len(cli_list_dow)

        _dow_str_fmt = wb.add_format({"font_name":"Calibri","font_size":11,"valign":"vcenter"})
        _dow_num_fmt = wb.add_format({"font_name":"Calibri","font_size":11,
                                       "num_format":"#,##0","align":"right","valign":"vcenter"})
        _dow_tbl_cols = [{"header":"Cliente","header_format":fmts["H"],"format":_dow_str_fmt}]
        for _d in _days_es:
            _dow_tbl_cols.append({"header":_d,"header_format":fmts["H"],"format":_dow_num_fmt})

        _dow_tbl_data = [[cli] + dow_by_cli[cli] for cli in cli_list_dow]

        if n_cli_dow > 0:
            ws_f.add_table(2, DOW_START_COL, 2 + n_cli_dow, DOW_START_COL + 7, {
                "name":"TablaDOWCliente","style":"Table Style Medium 2",
                "autofilter":False,"banded_rows":True,"first_column":False,
                "columns":_dow_tbl_cols,"data":_dow_tbl_data,
            })
        ws_f.set_column(DOW_START_COL, DOW_START_COL, 22)
        for _di in range(7):
            ws_f.set_column(DOW_START_COL+1+_di, DOW_START_COL+1+_di, 10)

        # Rangos para gráfico 1 (por fecha)
        ci_ff_xl = 1   # col B (Fecha)
        ci_ft_xl = 2   # col C (Total)
        ci_fe_xl = 3   # col D (Entregados)
        R1_f_xl  = 3   # primera fila de datos (0-indexed)
        RN_f_xl  = 2 + n_f

        chart_row_f = n_f + 5

        ch_f1 = _cs(wb, wb.add_chart({"type":"line"}))
        ch_f1.set_title({"name":"Pedidos por Fecha",**_tf()})
        ch_f1.set_x_axis({**_af("Fecha"),"text_axis":True})
        ch_f1.set_y_axis(_af("Pedidos"))
        ch_f1.add_series({"name":"Total",
                           "categories":["Fechas",R1_f_xl,ci_ff_xl,RN_f_xl,ci_ff_xl],
                           "values":    ["Fechas",R1_f_xl,ci_ft_xl,RN_f_xl,ci_ft_xl],
                           "line":{"color":CN,"width":2},
                           "marker":{"type":"circle","size":4,"fill":{"color":CN},"border":{"none":True}},
                           "data_labels":{"value":True,"font":{"name":"Arial","size":8,"bold":True}}})
        ch_f1.add_series({"name":"Entregados",
                           "categories":["Fechas",R1_f_xl,ci_ff_xl,RN_f_xl,ci_ff_xl],
                           "values":    ["Fechas",R1_f_xl,ci_fe_xl,RN_f_xl,ci_fe_xl],
                           "line":{"color":CB,"width":2},
                           "marker":{"type":"circle","size":4,"fill":{"color":CB},"border":{"none":True}},
                           "data_labels":{"value":True,"font":{"name":"Arial","size":8,"bold":True}}})
        ch_f1.set_size({"width":580,"height":280})
        ws_f.insert_chart(chart_row_f, 1, ch_f1)

        # Gráfico 2: Pedidos por Cliente por Día de la Semana — multi-línea
        # Una serie por cliente; categorías = encabezados de días en la tabla DOW
        _pal_dow = ["#1E6BB5","#10B981","#F59E0B","#8B5CF6","#EF4444",
                    "#4A9FE0","#06B6D4","#F97316","#84CC16","#EC4899"]
        ch_f2 = wb.add_chart({"type":"line"})
        ch_f2.set_style(10)
        ch_f2.set_chartarea({"border":{"none":True},"fill":{"color":"#FFFFFF"}})
        ch_f2.set_plotarea({"border":{"color":"#E2E8F0","width":0.75},"fill":{"color":"#FFFFFF"}})
        ch_f2.set_title({"name":"Pedidos por Cliente · Día de la Semana",**_tf()})
        ch_f2.set_x_axis({"name":"","major_gridlines":{"visible":True,"line":{"color":"#E8EDF2"}},
                           "num_font":{"size":9}})
        ch_f2.set_y_axis({"name":"","major_gridlines":{"visible":True,"line":{"color":"#E8EDF2"}},
                           "num_font":{"size":9}})
        ch_f2.set_legend({"position":"bottom","font":{"name":"Arial","size":9},"border":{"none":True}})
        for _si, _cli in enumerate(cli_list_dow):
            _ri = 3 + _si          # fila 0-indexed de datos del cliente en la hoja
            _col = _pal_dow[_si % len(_pal_dow)]
            ch_f2.add_series({
                "name":       ["Fechas", _ri, DOW_START_COL],
                "categories": ["Fechas", 2, DOW_START_COL+1, 2, DOW_START_COL+7],
                "values":     ["Fechas", _ri, DOW_START_COL+1, _ri, DOW_START_COL+7],
                "line":       {"color":_col,"width":2},
                "marker":     {"type":"circle","size":5,
                               "fill":{"color":_col},"border":{"none":True}},
                "data_labels":{"value":True,"font":{"name":"Arial","size":8,"bold":True}},
            })
        ch_f2.set_size({"width":540,"height":300})
        ws_f.insert_chart(chart_row_f, 11, ch_f2)

        # ══════════════════════════════════════════════════════════════
        # HOJA 4: ESTADO
        # ══════════════════════════════════════════════════════════════
        g_est = _agg_estado(df); n_e = len(g_est)
        ws_e, cm_e, _ = _write_module_sheet(wb, fmts, "Estado", g_est,
            f"Reporte por Estado ({n_e} estados)")
        R1_e, RN_e = 3, 2+n_e
        ci_ee = cm_e["Estado"]; ci_et = cm_e.get("Total",2); ci_eb = cm_e.get("Bultos",3)
        chart_row_e = n_e + 5

        ch_e1 = _cs(wb, wb.add_chart({"type":"doughnut"}), legend="right")
        ch_e1.set_title({"name":"Distribución por Estado",**_tf()})
        ch_e1.add_series({"name":"Pedidos","categories":["Estado",R1_e,ci_ee,RN_e,ci_ee],
                           "values":["Estado",R1_e,ci_et,RN_e,ci_et],
                           "data_labels":{"percentage":True,"category":True,"separator":"\n",
                                          "font":{"name":"Arial","size":9}},
                           "points":[{"fill":{"color":c}} for c in PAL[:n_e]]})
        ch_e1.set_hole_size(40); ch_e1.set_size({"width":440,"height":300})
        ws_e.insert_chart(chart_row_e, 1, ch_e1)

        ch_e2 = _cs(wb, wb.add_chart({"type":"column"}))
        ch_e2.set_title({"name":"Pedidos y Bultos por Estado",**_tf()})
        ch_e2.set_y_axis(_af("Cantidad"))
        ch_e2.add_series({"name":"Pedidos","categories":["Estado",R1_e,ci_ee,RN_e,ci_ee],
                           "values":["Estado",R1_e,ci_et,RN_e,ci_et],
                           "fill":{"color":CB},"border":{"none":True},
                           "data_labels":{"value":True,"font":{"name":"Arial","size":8,"bold":True}}})
        if ci_eb:
            ch_e2.add_series({"name":"Bultos","categories":["Estado",R1_e,ci_ee,RN_e,ci_ee],
                               "values":["Estado",R1_e,ci_eb,RN_e,ci_eb],
                               "fill":{"color":CT},"border":{"none":True},
                               "data_labels":{"value":True,"font":{"name":"Arial","size":8,"bold":True}}})
        ch_e2.set_size({"width":440,"height":280})
        ws_e.insert_chart(chart_row_e, 10, ch_e2)

        # ══════════════════════════════════════════════════════════════
        # HOJA 5: ZONAS
        # ══════════════════════════════════════════════════════════════
        g_zon = _agg_zonas(df).head(20); n_z = len(g_zon)
        ws_z, cm_z, _ = _write_module_sheet(wb, fmts, "Zonas", g_zon,
            f"Reporte por Zona (Top {n_z})")
        R1_z, RN_z = 3, 2+n_z
        ci_zk = cm_z["Zona"]; ci_ze = cm_z.get("Entregados"); ci_zt = cm_z.get("Total",2)
        rn8_z = min(2+8, 2+n_z); chart_row_z = n_z + 5

        ch_z1 = _cs(wb, wb.add_chart({"type":"bar"}))
        ch_z1.set_title({"name":"Pedidos Entregados por Zona",**_tf()})
        ch_z1.set_x_axis(_af("Pedidos"))
        ch_z1.set_y_axis({"name":"","line":{"none":True},"name_font":{"name":"Arial","size":9}})
        if ci_ze:
            ch_z1.add_series({"name":"Entregados","categories":["Zonas",R1_z,ci_zk,RN_z,ci_zk],
                               "values":["Zonas",R1_z,ci_ze,RN_z,ci_ze],
                               "fill":{"color":CB},"border":{"none":True},
                               "data_labels":{"value":True,"font":{"name":"Arial","size":8,"bold":True}}})
        ch_z1.set_size({"width":520,"height":max(260,n_z*32+80)})
        ws_z.insert_chart(chart_row_z, 1, ch_z1)

        ch_z2 = _cs(wb, wb.add_chart({"type":"doughnut"}), legend="right")
        ch_z2.set_title({"name":"Porcentaje de Participación",**_tf()})
        ch_z2.add_series({"name":"Pedidos","categories":["Zonas",R1_z,ci_zk,rn8_z,ci_zk],
                           "values":["Zonas",R1_z,ci_zt,rn8_z,ci_zt],
                           "data_labels":{"percentage":True,"category":True,"separator":"\n",
                                          "font":{"name":"Arial","size":9}},
                           "points":[{"fill":{"color":c}} for c in PAL[:8]]})
        ch_z2.set_hole_size(40); ch_z2.set_size({"width":440,"height":300})
        ws_z.insert_chart(chart_row_z, 7, ch_z2)

    except Exception as _agg_err:
        ws_aviso = wb.add_worksheet("Aviso")
        aviso_fmt = wb.add_format({"bold": True, "font_name": "Calibri",
                                    "font_size": 12, "font_color": "#B91C1C",
                                    "bg_color": "#FFF5F5", "border": 1,
                                    "text_wrap": True, "valign": "top"})
        ws_aviso.set_column(0, 0, 80)
        ws_aviso.set_row(0, 60)
        ws_aviso.write(
            0, 0,
            f"No se pudo generar alguna hoja de análisis.\n"
            f"Error: {_agg_err}\n\n"
            f"Verificá que el archivo tiene las columnas: "
            f"Cliente, Estado, Distribuidor, Pieza, Localidad Destino.",
            aviso_fmt,
        )

    wb.close()
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
#  MÓDULO 1 — CLIENTES
# ═══════════════════════════════════════════════════════════════════════════
def tab_clientes(df: pd.DataFrame):
    g     = _agg_clientes(df)
    n_cli = len(g)
    total = int(g["Total"].sum())
    n_ent = int(g["Entregados"].sum())
    n_bul = int(g["Bultos"].sum())
    pct   = n_ent/total*100 if total else 0
    top_c = g.iloc[0]["Cliente"] if n_cli else "—"
    top_n = int(g.iloc[0]["Total"]) if n_cli else 0
    top8_clientes = g.head(8)["Cliente"].tolist()
    bar_colors = [PAL[top8_clientes.index(c)%len(PAL)] if c in top8_clientes else "#CBD5E1"
                  for c in g.head(15)["Cliente"].tolist()]

    # Tarifa data
    tarifa_map = st.session_state.get("tarifa_map", {})
    g_tar      = _agg_tarifa(df, tarifa_map)
    total_cobrar = int(g_tar["Total a Cobrar"].sum()) if not g_tar.empty else 0

    with st.expander("INDICADORES", expanded=True):
        cols_n = 6 if tarifa_map else 5
        c1,c2,c3,c4,c5 = st.columns(5)
        with c1: kcard("Total Clientes",  fmt(n_cli),    "en el período",           "")
        with c2: kcard("Total Pedidos",   fmt(total),    f"{n_ent:,} entregados",   "g","up")
        with c3: kcard("% Efectividad",   f"{pct:.0f}%", f"{n_ent:,} de {total:,}","g")
        with c4: kcard("Cliente Top",     top_c,         f"{top_n:,} pedidos",      "p","",small_val=True)
        with c5:
            if tarifa_map:
                kcard("Total a Cobrar", f"${total_cobrar:,.0f}", f"{len(g_tar)} clientes tarifados","t")
            else:
                kcard("Bultos Totales",  fmt(n_bul),    "unidades despachadas",    "t")

    with st.expander("GRÁFICOS", expanded=True):
        top15 = g.head(15).sort_values("Entregados", ascending=True)
        bar_colors_asc = list(reversed(bar_colors[:len(top15)]))

        # Fila 1: 2 gráficos lado a lado
        ca, cb_ = st.columns([1.6, 1])
        with ca:
            co("Pedidos por Cliente")
            fig = go.Figure(go.Bar(
                y=top15["Cliente"], x=top15["Entregados"], orientation="h",
                marker_color=bar_colors_asc,
                text=top15["Entregados"], textposition="outside",
                hovertemplate="<b>%{y}</b><br>Entregados: <b>%{x:,}</b><extra></extra>"))
            theme(fig, h=max(300,len(top15)*46+80), mr=60)
            fig.update_yaxes(showgrid=False); pc(fig); cc()
        with cb_:
            co("Porcentaje de Participación")
            top8 = g.head(8)
            fig2 = go.Figure(go.Pie(
                labels=top8["Cliente"], values=top8["Total"], hole=0.52,
                marker=dict(colors=PAL[:len(top8)], line=dict(color="#FFF",width=2)),
                texttemplate="%{label}<br>%{percent:.0%}",
                textfont=dict(size=10), textposition="outside",
                hovertemplate="<b>%{label}</b><br>%{value:,} pedidos · %{percent:.0%}<extra></extra>"))
            fig2.add_annotation(text=f"<b>{fmt(total)}</b><br>pedidos",
                x=0.5,y=0.5,showarrow=False,font=dict(size=16,family="Inter",color=CN),align="center")
            theme(fig2, h=max(340,len(top15)*50+100))
            fig2.update_layout(showlegend=False)
            pc(fig2); cc()

        # Fila 2: Gráfico tarifa (debajo, mitad de pantalla)
        if tarifa_map and not g_tar.empty:
            half_col, _ = st.columns([1, 1])
            with half_col:
                g_tar_s = g_tar.sort_values("Total a Cobrar", ascending=True)
                co("Total Tarifa por Cliente")
                fig3 = go.Figure(go.Bar(
                    y=g_tar_s["Cliente"], x=g_tar_s["Total a Cobrar"],
                    orientation="h",
                    marker_color=CG,
                    text=[f"${v:,.0f}" for v in g_tar_s["Total a Cobrar"]],
                    textposition="outside",
                    hovertemplate="<b>%{y}</b><br>Total: <b>$%{x:,.0f}</b><extra></extra>"))
                theme(fig3, h=max(280, len(g_tar_s)*52+80), ml=10, mr=80)
                fig3.update_xaxes(tickprefix="$", tickformat=",", showgrid=True, title_text="")
                fig3.update_yaxes(showgrid=False, title_text="")
                fig3.update_layout(showlegend=False)
                pc(fig3); cc()

    with st.expander("TABLA DE DATOS", expanded=False):
        co("Tabla por Cliente")
        # Agregar columna tarifa si hay datos
        g_show = g[["Cliente","Total","Entregados","Pendientes","Bultos","% Efectividad"]].copy()
        if tarifa_map and not g_tar.empty:
            tar_col = g_tar.set_index("Cliente")["Total a Cobrar"]
            g_show["Tarifa Cliente"] = g_show["Cliente"].map(tar_col).fillna(0).astype(int)
            # Formatear como pesos ARS sin decimales para display
            g_display = g_show.copy()
            g_display["Tarifa Cliente"] = g_display["Tarifa Cliente"].apply(
                lambda v: f"$ {int(v):,}".replace(",",".") if v > 0 else "—")
            plain_table(g_display, ["Total","Entregados","Pendientes","Bultos"], ["% Efectividad"])
        else:
            plain_table(g_show, ["Total","Entregados","Pendientes","Bultos"], ["% Efectividad"])
        cc()


# ═══════════════════════════════════════════════════════════════════════════
#  MÓDULO 2 — DISTRIBUIDORES
# ═══════════════════════════════════════════════════════════════════════════
def tab_distribuidores(df: pd.DataFrame):
    g     = _agg_distrib(df)
    n_d   = len(g)
    total = int(g["Total"].sum())
    n_ent = int(g["Entregados"].sum())
    pct   = n_ent/total*100 if total else 0
    best  = g.loc[g["% Efectividad"].idxmax(),"Distribuidor"] if n_d else "—"
    best_p= float(g["% Efectividad"].max()) if n_d else 0
    prom  = round(total/n_d,1) if n_d else 0

    with st.expander("INDICADORES", expanded=True):
        c1,c2,c3,c4,c5 = st.columns(5)
        with c1: kcard("Distribuidores",         fmt(n_d),       "en operación",             "")
        with c2: kcard("Total Pedidos",          fmt(total),     f"{n_ent:,} entregados",    "g","up")
        with c3: kcard("% Efectividad General",  f"{pct:.0f}%",  f"{n_ent:,} de {total:,}", "g")
        with c4: kcard("Mejor Efectividad",      best,           f"{best_p:.0f}% efectividad","p","",small_val=True)
        with c5: kcard("Prom. Pedidos / Driver", str(prom),      "pedidos por distribuidor", "t")

    with st.expander("GRÁFICOS", expanded=True):
        g_asc = g.sort_values("Entregados", ascending=True)
        ca, cb_ = st.columns([1.3,1])
        with ca:
            co("Pedidos Entregados por Distribuidor")
            fig = go.Figure(go.Bar(
                y=g_asc["Distribuidor"], x=g_asc["Entregados"], orientation="h",
                marker_color=CB,
                text=g_asc["Entregados"], textposition="outside",
                hovertemplate="<b>%{y}</b><br>Entregados: <b>%{x:,}</b><extra></extra>"))
            theme(fig, h=max(300,n_d*52+80), mr=60)
            fig.update_yaxes(showgrid=False); pc(fig); cc()
        with cb_:
            co("% Efectividad por Distribuidor")
            colors = [CG if v>=85 else CA if v>=65 else CR for v in g["% Efectividad"]]
            fig2 = go.Figure(go.Bar(
                x=g["Distribuidor"], y=g["% Efectividad"],
                marker_color=colors, width=0.35,
                text=g["% Efectividad"].astype(str)+"%", textposition="outside",
                hovertemplate="<b>%{x}</b><br>Efectividad: <b>%{y:.1f}%</b><extra></extra>"))
            fig2.add_hline(y=85, line_dash="dot", line_color=CG,
                           annotation_text="Meta 85%", annotation_font_size=10)
            fig2.update_layout(yaxis_range=[0,118], bargap=0.6)
            theme(fig2, h=340); pc(fig2); cc()

    with st.expander("TABLA DE DATOS", expanded=False):
        co("Tabla por Distribuidor")
        plain_table(g[["Distribuidor","Total","Entregados","Pendientes","Bultos","% Efectividad"]],
                    ["Total","Entregados","Pendientes","Bultos"],["% Efectividad"])
        cc()


# ═══════════════════════════════════════════════════════════════════════════
#  MÓDULO 3 — FECHAS
# ═══════════════════════════════════════════════════════════════════════════
def tab_fecha(df: pd.DataFrame):
    g      = _agg_fecha(df)
    g_dow  = _agg_dow(df)
    total  = int(g["Total"].sum())
    n_ent  = int(g["Entregados"].sum())
    pct    = n_ent/total*100 if total else 0
    pico   = int(g["Total"].max()) if len(g) else 0
    prom_d = round(float(g["Total"].mean()),1) if len(g) else 0
    fechas_cortas = [fecha_corta(f) for f in g["Fecha"].tolist()]

    # Mostrar columna de fecha usada
    fcol = get_fecha_col(df)

    with st.expander("INDICADORES", expanded=True):
        c1,c2,c3,c4,c5 = st.columns(5)
        with c1: kcard("Días con Actividad", fmt(len(g)),   "días del período",         "")
        with c2: kcard("Total Pedidos",      fmt(total),    f"{n_ent:,} entregados",    "g","up")
        with c3: kcard("% Efectividad",      f"{pct:.0f}%", f"{n_ent:,} de {total:,}", "g")
        with c4: kcard("Pico Máximo",        fmt(pico),     "pedidos en un día",        "a")
        with c5: kcard("Promedio Diario",    str(prom_d),   "pedidos por día",          "t")
        if fcol:
            st.caption(f"📅 Usando columna de fecha: **{fcol}**")

    with st.expander("GRÁFICOS", expanded=True):
        co("Pedidos por Fecha")
        fig = go.Figure(go.Scatter(
            x=fechas_cortas, y=g["Total"].tolist(),
            mode="lines+markers+text",
            line=dict(color=CB, width=2.5, shape="spline"),
            marker=dict(size=7, color=CB, line=dict(color="#fff",width=1.5)),
            fill="tozeroy", fillcolor="rgba(30,107,181,0.10)",
            text=g["Total"].tolist(), textposition="top center",
            textfont=dict(size=10, color=CB),
            hovertemplate="<b>%{x}</b><br>Pedidos: <b>%{y:,}</b><extra></extra>",
            name="Pedidos"))
        fig.update_layout(hovermode="x unified", showlegend=False,
                          xaxis=dict(tickangle=-40, tickfont=dict(size=11)))
        fig.update_xaxes(title_text=""); fig.update_yaxes(title_text="")
        theme(fig, h=400); pc(fig); cc()

        co("Pedidos por Día de la Semana")
        fig2 = go.Figure(go.Scatter(
            x=g_dow["Dia"].tolist(), y=g_dow["Total"].tolist(),
            mode="lines+markers+text",
            line=dict(color=CG, width=2.5),
            marker=dict(size=9, color=CG, line=dict(color="#fff",width=1.5)),
            fill="tozeroy", fillcolor="rgba(16,185,129,0.10)",
            text=g_dow["Total"].tolist(), textposition="top center",
            textfont=dict(size=11, color=CG),
            hovertemplate="<b>%{x}</b><br>Pedidos: <b>%{y:,}</b><extra></extra>"))
        fig2.update_layout(hovermode="x unified", showlegend=False)
        fig2.update_xaxes(title_text=""); fig2.update_yaxes(title_text="")
        theme(fig2, h=320); pc(fig2); cc()

        # ── Gráfico: Pedidos por Cliente por Día de la Semana ─────────
        dow_data, dias = _agg_dow_by_client(df)
        if dow_data:
            co("Pedidos por Cliente · Día de la Semana")
            fig3 = go.Figure()
            for i, (cli, vals) in enumerate(dow_data.items()):
                color = PAL[i % len(PAL)]
                r, g_hex, b_hex = (int(color[1:3],16),
                                   int(color[3:5],16),
                                   int(color[5:7],16))
                fig3.add_trace(go.Scatter(
                    x=dias, y=vals,
                    name=cli,
                    mode="lines+markers+text",
                    line=dict(color=color, width=2.2),
                    marker=dict(size=8, color=color,
                                line=dict(color="#fff", width=1.5)),
                    text=vals, textposition="top center",
                    textfont=dict(size=9, color=color),
                    hovertemplate=f"<b>{cli}</b><br>%{{x}}: <b>%{{y:,}}</b><extra></extra>",
                    fill="tozeroy",
                    fillcolor=f"rgba({r},{g_hex},{b_hex},0.06)",
                ))
            fig3.update_layout(
                hovermode="x unified",
                showlegend=True,
                legend=dict(orientation="h", yanchor="bottom", y=-0.28,
                            xanchor="center", x=0.5,
                            font=dict(size=11), bgcolor="rgba(0,0,0,0)"),
            )
            fig3.update_xaxes(title_text="")
            fig3.update_yaxes(title_text="")
            theme(fig3, h=360); pc(fig3); cc()

    with st.expander("TABLA DE DATOS", expanded=False):
        co("Tabla de Actividad Diaria")
        plain_table(g[["Fecha","Total","Entregados","Pendientes","Bultos"]],
                    ["Total","Entregados","Pendientes","Bultos"])
        cc()


# ═══════════════════════════════════════════════════════════════════════════
#  MÓDULO 4 — ESTADO
# ═══════════════════════════════════════════════════════════════════════════
def tab_estado(df: pd.DataFrame):
    g       = _agg_estado(df)
    total   = int(g["Total"].sum())
    n_ent   = int(is_ent(safe(df,COL_ESTADO)).sum())
    n_pend  = total - n_ent
    pct     = n_ent/total*100 if total else 0
    n_est   = len(g)

    with st.expander("INDICADORES", expanded=True):
        # Siempre 3 tarjetas fijas + tarjetas por estado
        c_fixed = st.columns(3 + min(n_est, 3))
        with c_fixed[0]: kcard("Total Pedidos",    fmt(total),  f"{pct:.0f}% entregados",  "g","up")
        with c_fixed[1]: kcard("Pedidos Entregados",fmt(n_ent), "pedidos entregados",       "g")
        with c_fixed[2]: kcard("Pedidos Pendientes",fmt(n_pend),"pedidos sin entregar",     "r")
        c_map = ["a","p","t"]
        for i, row in g.head(3).iterrows():
            pct_e = row["Total"]/total*100 if total else 0
            with c_fixed[3+i]:
                kcard(row["Estado"], fmt(int(row["Total"])),
                      f"{pct_e:.0f}% del total", c_map[i%3])

    with st.expander("GRÁFICOS", expanded=True):
        ca, cb_ = st.columns([1,1.4])
        with ca:
            co("Distribución por Estado")
            fig = go.Figure(go.Pie(
                labels=g["Estado"], values=g["Total"], hole=0.55,
                marker=dict(colors=PAL[:n_est], line=dict(color="#FFF",width=2)),
                texttemplate="%{label}<br>%{percent:.0%}",
                textfont=dict(size=11),
                hovertemplate="<b>%{label}</b><br>%{value:,} pedidos — %{percent:.0%}<extra></extra>"))
            fig.add_annotation(text=f"<b>{fmt(total)}</b><br>pedidos",
                x=0.5,y=0.5,showarrow=False,font=dict(size=16,family="Inter",color=CN),align="center")
            theme(fig,h=380); pc(fig); cc()
        with cb_:
            co("Pedidos y Bultos por Estado")
            fig2 = go.Figure()
            fig2.add_bar(x=g["Estado"], y=g["Total"], name="Pedidos",
                         marker_color=CB, width=0.3,
                         text=g["Total"], textposition="outside",
                         hovertemplate="<b>%{x}</b><br>Pedidos: <b>%{y:,}</b><extra></extra>")
            fig2.add_bar(x=g["Estado"], y=g["Bultos"], name="Bultos",
                         marker_color=CT, width=0.3,
                         text=g["Bultos"], textposition="outside",
                         hovertemplate="<b>%{x}</b><br>Bultos: <b>%{y:,}</b><extra></extra>")
            fig2.update_layout(barmode="group", bargap=0.5)
            theme(fig2,h=380); pc(fig2); cc()

    with st.expander("TABLA DE DATOS", expanded=False):
        co("Tabla por Estado")
        plain_table(g[["Estado","Total","Bultos","Clientes","Distribuidores"]],
                    ["Total","Bultos","Clientes","Distribuidores"])
        cc()


# ═══════════════════════════════════════════════════════════════════════════
#  MÓDULO 5 — ZONAS
# ═══════════════════════════════════════════════════════════════════════════
def tab_zonas(df: pd.DataFrame):
    g     = _agg_zonas(df)
    n_z   = len(g)
    total = int(g["Total"].sum())
    n_ent = int(g["Entregados"].sum())
    pct   = n_ent/total*100 if total else 0
    n_bul = int(g["Bultos"].sum())
    top_z = g.iloc[0]["Zona"] if n_z else "—"
    top_n = int(g.iloc[0]["Total"]) if n_z else 0
    top8_zonas  = g.head(8)["Zona"].tolist()
    top15_zonas = g.head(15)["Zona"].tolist()
    bar_colors_z     = [PAL[top8_zonas.index(z)%len(PAL)] if z in top8_zonas else "#CBD5E1"
                        for z in top15_zonas]
    bar_colors_z_asc = list(reversed(bar_colors_z))

    with st.expander("INDICADORES", expanded=True):
        c1,c2,c3,c4,c5 = st.columns(5)
        with c1: kcard("Zonas Totales",   fmt(n_z),    "en el período",            "")
        with c2: kcard("Total Pedidos",   fmt(total),  f"{n_ent:,} entregados",    "g","up")
        with c3: kcard("% Efectividad",   f"{pct:.0f}%",f"{n_ent:,} de {total:,}","g")
        with c4: kcard("Zona Líder",      top_z,       f"{top_n:,} pedidos",       "p","",small_val=True)
        with c5: kcard("Bultos Totales",  fmt(n_bul),  "unidades despachadas",     "t")

    with st.expander("GRÁFICOS", expanded=True):
        top15 = g.head(15).sort_values("Entregados", ascending=True)
        ca, cb_ = st.columns([1.5,1])
        with ca:
            co("Pedidos Entregados por Zona")
            fig = go.Figure(go.Bar(
                y=top15["Zona"], x=top15["Entregados"], orientation="h",
                marker_color=bar_colors_z_asc,
                text=top15["Entregados"], textposition="outside",
                hovertemplate="<b>%{y}</b><br>Entregados: <b>%{x:,}</b><extra></extra>"))
            theme(fig, h=max(320,len(top15)*46+80), mr=60)
            fig.update_yaxes(showgrid=False); pc(fig); cc()
        with cb_:
            co("Porcentaje de Participación")
            top8 = g.head(8)
            fig2 = go.Figure(go.Pie(
                labels=top8["Zona"], values=top8["Total"], hole=0.52,
                marker=dict(colors=PAL[:len(top8)], line=dict(color="#FFF",width=2)),
                texttemplate="%{label}<br>%{percent:.0%}",
                textfont=dict(size=10), textposition="outside",
                hovertemplate="<b>%{label}</b><br>%{value:,} pedidos · %{percent:.0%}<extra></extra>"))
            fig2.add_annotation(text=f"<b>{fmt(total)}</b><br>pedidos",
                x=0.5,y=0.5,showarrow=False,font=dict(size=16,family="Inter",color=CN),align="center")
            theme(fig2, h=max(320,len(top15)*46+80)); pc(fig2); cc()

    with st.expander("TABLA DE DATOS", expanded=False):
        co("Tabla por Zona")
        plain_table(g[["Zona","Total","Entregados","Pendientes","Bultos","Clientes","Distribuidores","% Efectividad"]],
                    ["Total","Entregados","Pendientes","Bultos","Clientes","Distribuidores"],["% Efectividad"])
        cc()



# ═══════════════════════════════════════════════════════════════════════════
#  DASHBOARD PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════
def render_dashboard(df: pd.DataFrame, fname: str):
    # Header
    c_logo, c_btns = st.columns([3, 1])
    with c_logo:
        st.markdown(f"""
        <div style='padding-top:4px;padding-bottom:8px'>
          <div style='display:flex;align-items:center;gap:10px'>
            <svg width="32" height="32" viewBox="0 0 64 64">
              <rect x="4"  y="38" width="14" height="20" rx="3" fill="#1E6BB5"/>
              <rect x="24" y="22" width="14" height="36" rx="3" fill="#1E6BB5"/>
              <rect x="44" y="10" width="14" height="48" rx="3" fill="#0F4C8A"/>
              <line x1="2" y1="61" x2="62" y2="61" stroke="#0F1C2E" stroke-width="2" stroke-linecap="round"/>
            </svg>
            <div>
              <div style='font-size:20px;font-weight:900;color:#0F1C2E;line-height:1;letter-spacing:-.5px'>
                Flash Expres <span style='color:#1E6BB5'>- Analytics</span>
              </div>
              <div style='font-size:11px;color:#94A3B8;margin-top:1px'>
                {fname} &nbsp;·&nbsp; {len(df):,} pedidos analizados
              </div>
            </div>
          </div>
        </div>""", unsafe_allow_html=True)
    with c_btns:
        st.markdown("<div style='padding-top:6px'></div>", unsafe_allow_html=True)
        b1, b2 = st.columns(2)
        with b1:
            if st.button("← Nuevo archivo", use_container_width=True):
                for k in ["df","fname","confirmed","stats","tarifa_rows","tarifa_map"]:
                    st.session_state.pop(k, None)
                st.rerun()
        with b2:
            excel_bytes = build_full_excel(df, st.session_state.get('tarifa_map', {}))
            import datetime
            fname_out = f"Flash_Analytics_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
            st.download_button("⬇ Descargar Excel", data=excel_bytes,
                               file_name=fname_out,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

    st.markdown("<hr style='margin:0 0 12px;border:none;border-top:1px solid #E2E8F0'>",
                unsafe_allow_html=True)

    tabs = st.tabs(["  Clientes","  Distribuidores","  Fechas","  Estado","  Zonas"])
    with tabs[0]: tab_clientes(df)
    with tabs[1]: tab_distribuidores(df)
    with tabs[2]: tab_fecha(df)
    with tabs[3]: tab_estado(df)
    with tabs[4]: tab_zonas(df)


# ===================================================
#  PANTALLA HOME
# ===================================================
# ===================================================
#  PANTALLA HOME — SVG gráficos de fondo + upload
# ===================================================
# ===================================================
#  PANTALLA HOME
# ===================================================
def render_home():
    import streamlit.components.v1 as components


    # CSS upload widget — oculta TODO el texto nativo del boton
    st.markdown("""
    <style>
    [data-testid="stFileUploaderDropzone"] button {
        background:#1E6BB5!important; border:none!important;
        border-radius:8px!important; font-size:0!important;
        color:transparent!important; position:relative!important;
        min-width:140px!important; height:38px!important }
    [data-testid="stFileUploaderDropzone"] button:hover { background:#0F4C8A!important }
    [data-testid="stFileUploaderDropzone"] button * {
        font-size:0!important; color:transparent!important;
        visibility:hidden!important; width:0!important; overflow:hidden!important }
    [data-testid="stFileUploaderDropzone"] button::after {
        content:"Cargar Archivo" !important;
        font-size:14px!important; font-weight:700!important; color:#ffffff!important;
        visibility:visible!important; display:block!important;
        position:absolute!important; top:50%!important; left:50%!important;
        transform:translate(-50%,-50%)!important; white-space:nowrap!important }
    [data-testid="stFileUploaderDropzoneInstructions"] div span:first-child::before {
        content:"Arrastrá tu archivo aquí"; display:block; font-weight:600; color:#475569 }
    [data-testid="stFileUploaderDropzoneInstructions"] div span:first-child { display:none }
    [data-testid="stFileUploaderDropzoneInstructions"] small::before {
        content:"Límite 200MB por archivo"; display:block; font-size:12px; color:#94A3B8 }
    [data-testid="stFileUploaderDropzoneInstructions"] small { display:none }
    </style>""", unsafe_allow_html=True)

    st.markdown("<div style='height:60px'></div>", unsafe_allow_html=True)
    _, col_c, _ = st.columns([1, 2, 1])
    with col_c:
        st.markdown("""
        <div style='text-align:center;margin-bottom:28px;position:relative;z-index:1'>
            <svg width="60" height="60" viewBox="0 0 64 64" style="display:block;margin:0 auto 14px">
              <rect x="4"  y="38" width="14" height="20" rx="3" fill="#1E6BB5"/>
              <rect x="24" y="22" width="14" height="36" rx="3" fill="#1E6BB5"/>
              <rect x="44" y="10" width="14" height="48" rx="3" fill="#0F4C8A"/>
              <line x1="2" y1="61" x2="62" y2="61" stroke="#0F1C2E" stroke-width="2" stroke-linecap="round"/>
            </svg>
            <h1 style='font-size:28px;font-weight:900;color:#0F1C2E;margin:0 0 6px;letter-spacing:-1px'>
                Flash Expres<span style='color:#1E6BB5'> - Analytics</span>
            </h1>
            <p style='font-size:14px;color:#64748B;margin:0'>Plataforma de Análisis Logístico</p>
        </div>""", unsafe_allow_html=True)
        st.markdown("<div style='position:relative;z-index:1'>", unsafe_allow_html=True)
        uploaded = st.file_uploader("Subir archivo", type=["csv","xlsx","xls"],
                                    label_visibility="collapsed", key="file_upload")
        st.markdown("</div>", unsafe_allow_html=True)

    return uploaded


# ═══════════════════════════════════════════════════════════════════
#  PANTALLA DE CONFIRMACIÓN — diseño profesional centrado
# ═══════════════════════════════════════════════════════════════════
def render_popup(stats: dict, fname: str):
    """Pantalla de confirmación. Card CSS en columna central. Tarifas lazy."""
    ok         = stats["valid"]
    miss_val   = ", ".join(stats["missing_cols"]) if stats["missing_cols"] else "Ninguna"
    status_lbl = "Archivo cargado correctamente" if ok else "Se detectaron problemas en el archivo"
    status_clr = "#1E5FA8" if ok else "#D97706"
    msg        = ("Revisá los datos del archivo y configurá las tarifas si corresponde."
                  if ok else "Se detectaron problemas. Podés continuar de todas formas o cancelar.")

    # ── CSS + JS ─────────────────────────────────────────────────────────────
    st.markdown(f"""
<style>
/* ══ CARD — aplica SOLO al stVerticalBlock del 2º stColumn externo ══════ */
[data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:nth-child(2)
    > [data-testid="stVerticalBlock"] {{
    background    : #FFFFFF;
    border        : 1px solid #C8D6E8;
    border-radius : 14px;
    padding       : 20px 28px 24px !important;
    box-shadow    : 0 6px 28px rgba(15,28,46,0.14),
                    0 2px 8px  rgba(15,28,46,0.08),
                    0 0 0 1px  rgba(15,28,46,0.04);
    margin-top    : 4px;
}}
[data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:nth-child(2)
    > [data-testid="stVerticalBlock"] > div {{ gap:0 !important; }}

/* ══ RESET: cualquier stVerticalBlock DENTRO del card (filas de tarifa,
   fila de botones, etc.) NO debe heredar el estilo del card ══════════ */
[data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:nth-child(2)
    > [data-testid="stVerticalBlock"]
    [data-testid="stHorizontalBlock"]
    [data-testid="stColumn"]
    > [data-testid="stVerticalBlock"] {{
    background    : transparent !important;
    border        : none !important;
    box-shadow    : none !important;
    padding       : 0 !important;
    margin-top    : 0 !important;
    border-radius : 0 !important;
}}

/* ══ Labels ═══════════════════════════════════════════════════════════ */
div[data-testid="InputInstructions"] {{ display:none !important; }}
div[data-testid="stMultiSelect"] > label,
div[data-testid="stTextInput"]   > label {{
    font-size:11px !important; font-weight:600 !important;
    color:#374151 !important; margin-bottom:2px !important;
}}

/* ══ CAMPO TARIFA ══════════════════════════════════════════════════════
   Streamlit text_input: el borde visible está en stTextInputRootElement.
   Apuntamos también a [data-baseweb="input"] como fallback.
   ════════════════════════════════════════════════════════════════════ */
div[data-testid="stTextInput"] [data-testid="stTextInputRootElement"],
div[data-testid="stTextInput"] [data-baseweb="input"] {{
    background-color : #FFFFFF !important;
    border           : 1px solid #D1D5DB !important;
    border-radius    : 6px !important;
    box-shadow       : none !important;
    min-height       : 42px !important;
    height           : 42px !important;
    padding          : 0 !important;
    display          : flex !important;
    align-items      : center !important;
    outline          : none !important;
}}
div[data-testid="stTextInput"] [data-testid="stTextInputRootElement"]:focus-within,
div[data-testid="stTextInput"] [data-baseweb="input"]:focus-within {{
    border-color : #6B9FD4 !important;
    box-shadow   : 0 0 0 2px rgba(30,95,168,0.12) !important;
}}
div[data-testid="stTextInput"] input {{
    background   : transparent !important;
    border       : none !important;
    box-shadow   : none !important;
    outline      : none !important;
    font-size    : 14px !important;
    color        : #111827 !important;
    padding      : 0 12px !important;
    height       : 40px !important;
    width        : 100% !important;
}}
div[data-testid="stTextInput"] input::placeholder {{
    color:#9CA3AF !important; opacity:1 !important;
}}

/* ══ Botón ✕ ══════════════════════════════════════════════════════════ */
button[aria-label="✕"] {{
    background-color:#FFFFFF !important; color:#9CA3AF !important;
    border:1px solid #E5E7EB !important; border-radius:6px !important;
    font-size:14px !important; font-weight:700 !important;
    height:42px !important; width:100% !important;
    padding:0 !important; outline:none !important; box-shadow:none !important;
}}
button[aria-label="✕"]:hover {{ background-color:#FEF2F2 !important; color:#DC2626 !important; border-color:#FECACA !important; }}
button[aria-label="✕"]:focus, button[aria-label="✕"]:focus-visible {{ outline:none !important; box-shadow:none !important; }}

/* ══ Botón + Agregar cliente — igual al campo de clientes ═══════════════ */
button[aria-label="+ Agregar cliente"] {{
    background-color : #FFFFFF !important;
    color            : #9CA3AF !important;
    border           : 1px solid #D1D5DB !important;
    border-radius    : 6px !important;
    font-weight      : 400 !important;
    font-size        : 14px !important;
    height           : 42px !important;
    width            : 100% !important;
    padding-left     : 12px !important;
    text-align       : left !important;
    outline          : none !important;
    box-shadow       : none !important;
}}
button[aria-label="+ Agregar cliente"]:hover {{
    background-color : #F9FAFB !important; border-color : #9CA3AF !important;
    color            : #6B7280 !important;
}}
button[aria-label="+ Agregar cliente"]:focus,
button[aria-label="+ Agregar cliente"]:focus-visible {{
    outline : none !important; box-shadow : none !important;
}}

/* ══ BOTONES CONTINUAR / CANCELAR ════════════════════════════════════
   Continuar = type="primary"  → kind="primary"  → AZUL
   Cancelar  = type="secondary"→ kind="secondary" → PLOMO
   React gestiona el atributo kind, nunca lo borra → CSS siempre aplica.
   ════════════════════════════════════════════════════════════════════ */

/* ── FORMA compartida (aplica a ambos) ── */
[data-testid="stButton"] button[kind="primary"],
[data-testid="stButton"] button[kind="secondary"] {{
    border-radius   : 8px !important;
    font-family     : inherit !important;
    font-weight     : 700 !important;
    font-size       : 15px !important;
    height          : 46px !important;
    min-height      : 46px !important;
    width           : 100% !important;
    border          : none !important;
    outline         : none !important;
    box-shadow      : none !important;
    padding         : 0 24px !important;
    cursor          : pointer !important;
    display         : flex !important;
    align-items     : center !important;
    justify-content : center !important;
    box-sizing      : border-box !important;
    color           : #FFFFFF !important;
    transition      : background-color 0.15s !important;
}}

/* ── COLOR Continuar — azul ── */
[data-testid="stButton"] button[kind="primary"] {{
    background-color : #1E5FA8 !important;
}}
[data-testid="stButton"] button[kind="primary"]:hover {{
    background-color : #164f8e !important;
}}

/* ── COLOR Cancelar — plomo ── */
[data-testid="stButton"] button[kind="secondary"] {{
    background-color : #6B7280 !important;
    color            : #FFFFFF !important;
}}
[data-testid="stButton"] button[kind="secondary"]:hover {{
    background-color : #4B5563 !important;
}}

/* ── Eliminar el "cuadro" del wrapper secondary de Streamlit ── */
[data-testid="stBaseButton-secondary"] {{
    background  : transparent !important;
    border      : none !important;
    box-shadow  : none !important;
    padding     : 0 !important;
}}

/* ── Quitar focus ring en ambos ── */
[data-testid="stButton"] button[kind="primary"]:focus,
[data-testid="stButton"] button[kind="primary"]:focus-visible,
[data-testid="stButton"] button[kind="secondary"]:focus,
[data-testid="stButton"] button[kind="secondary"]:focus-visible {{
    outline    : none !important;
    box-shadow : none !important;
}}

/* ══ OVERRIDE "+ Agregar cliente" — mayor especificidad (0,2,1) que kind=secondary ══ */
[data-testid="stButton"] button[aria-label="+ Agregar cliente"] {{
    background-color : #FFFFFF !important;
    color            : #9CA3AF !important;
    border           : 1px solid #D1D5DB !important;
    border-radius    : 6px !important;
    font-weight      : 400 !important;
    font-size        : 14px !important;
    height           : 42px !important;
    width            : 100% !important;
    padding-left     : 12px !important;
    text-align       : left !important;
    outline          : none !important;
    box-shadow       : none !important;
}}
[data-testid="stButton"] button[aria-label="+ Agregar cliente"]:hover {{
    background-color : #F9FAFB !important;
    border-color     : #9CA3AF !important;
    color            : #6B7280 !important;
}}
[data-testid="stButton"] button[aria-label="+ Agregar cliente"]:focus,
[data-testid="stButton"] button[aria-label="+ Agregar cliente"]:focus-visible {{
    outline : none !important; box-shadow : none !important;
}}
</style>
<script>
/* ── "+ Agregar cliente" → fondo blanco igual al campo de clientes ── */
(function fixAgregar() {{
    function styleAgregar() {{
        document.querySelectorAll('button').forEach(function(b) {{
            if ((b.innerText||b.textContent||'').trim() === '+ Agregar cliente') {{
                b.style.setProperty('background-color','#FFFFFF','important');
                b.style.setProperty('color','#9CA3AF','important');
                b.style.setProperty('border','1px solid #D1D5DB','important');
                b.style.setProperty('border-radius','6px','important');
                b.style.setProperty('font-size','14px','important');
                b.style.setProperty('height','42px','important');
                b.style.setProperty('text-align','left','important');
                b.style.setProperty('padding-left','12px','important');
                b.style.setProperty('box-shadow','none','important');
                b.style.setProperty('outline','none','important');
                if (!b._agr) {{
                    b._agr = true;
                    b.addEventListener('mouseenter', function() {{
                        this.style.setProperty('background-color','#F9FAFB','important');
                    }});
                    b.addEventListener('mouseleave', function() {{
                        this.style.setProperty('background-color','#FFFFFF','important');
                    }});
                }}
            }}
        }});
    }}
    styleAgregar();
    [100,300,700,1500].forEach(function(ms){{ setTimeout(styleAgregar,ms); }});
    new MutationObserver(function(muts) {{
        var found = muts.some(function(m){{ return m.addedNodes.length>0; }});
        if (found) setTimeout(styleAgregar,10);
    }}).observe(document.body,{{childList:true,subtree:true}});
}})();

/* ── Moneda en campo Tarifa ─────────────────────────────────────────── */
function setupCurrency() {{
    document.querySelectorAll('[data-testid="stTextInput"] input').forEach(function(inp) {{
        if (inp._cs) return; inp._cs = true;
        inp.addEventListener('blur',  function() {{
            var r = this.value.replace(/[^0-9]/g, '');
            if (r) this.value = parseInt(r, 10).toLocaleString('en-US');
        }});
        inp.addEventListener('focus', function() {{ this.value = this.value.replace(/,/g, ''); }});
    }});
}}
[200, 700, 1500].forEach(function(ms) {{ setTimeout(setupCurrency, ms); }});

/* ══ CERRAR DROPDOWN MULTISELECT ══════════════════════════════════════
   Estrategia: BaseWebUI agrega [data-baseweb="tag"] al DOM en el momento
   exacto en que la selección queda registrada.
   → Detectar esa inserción vía MutationObserver.
   → Encontrar el <input> DENTRO del mismo [data-baseweb="select"].
   → Despachar Escape sobre ese input: BaseWebUI tiene onKeyDown que
     llama closeMenu() cuando recibe Escape en su propio input.
   Verificar que el listbox siga en el DOM para no cerrar en carga inicial.
   ══════════════════════════════════════════════════════════════════════ */
new MutationObserver(function(muts) {{
    muts.forEach(function(m) {{
        m.addedNodes.forEach(function(n) {{
            if (n.nodeType !== 1) return;

            /* ¿Es este nodo (o contiene) un chip de selección? */
            var tag = null;
            if (n.getAttribute && n.getAttribute('data-baseweb') === 'tag') {{
                tag = n;
            }} else if (n.querySelector) {{
                tag = n.querySelector('[data-baseweb="tag"]');
            }}
            if (!tag) return;

            /* Solo actuar si hay un listbox abierto en el DOM */
            if (!document.querySelector('[role="listbox"]')) return;

            /* Encontrar el select contenedor del chip */
            var sel = tag.closest ? tag.closest('[data-baseweb="select"]') : null;
            if (!sel) {{
                /* fallback: primer select con listbox activo */
                sel = document.querySelector('[data-baseweb="select"]');
            }}
            if (!sel) return;

            var inp = sel.querySelector('input');
            if (!inp) return;

            /* Esperar 60ms para que BaseWebUI termine de registrar la selección,
               luego enviar Escape directamente al input — BaseWebUI lo recibe
               en su onKeyDown y ejecuta closeMenu() */
            setTimeout(function() {{
                if (!document.querySelector('[role="listbox"]')) return; /* ya cerró */
                inp.dispatchEvent(new KeyboardEvent('keydown', {{
                    key: 'Escape', code: 'Escape',
                    keyCode: 27, which: 27,
                    bubbles: true, cancelable: true
                }}));
            }}, 60);
        }});
    }});
}}).observe(document.body, {{childList: true, subtree: true}});
</script>
""", unsafe_allow_html=True)

    # ── Layout: 3 columnas; la central recibe el estilo de card vía CSS ──────
    _, col_c, _ = st.columns([1, 3, 1])

    with col_c:

        # ─── ENCABEZADO ─────────────────────────────────────────────────────
        _icon_bg  = "#1E5FA8" if ok else "#DC2626"
        _icon_svg = (
            '<circle cx="14" cy="14" r="14" fill="{c}"/>'
            '<polyline points="8,14 12,18.5 20,9.5" stroke="white" stroke-width="2" '
            'stroke-linecap="round" stroke-linejoin="round" fill="none"/>'
        ).format(c=_icon_bg) if ok else (
            '<circle cx="14" cy="14" r="14" fill="{c}"/>'
            '<line x1="9" y1="9" x2="19" y2="19" stroke="white" stroke-width="2" stroke-linecap="round"/>'
            '<line x1="19" y1="9" x2="9" y2="19" stroke="white" stroke-width="2" stroke-linecap="round"/>'
        ).format(c=_icon_bg)

        st.markdown(f"""
<div style="display:flex;align-items:center;gap:12px;
            padding-bottom:14px;border-bottom:1px solid #E8EDF3;margin-bottom:14px">
  <svg width="28" height="28" viewBox="0 0 28 28" style="flex-shrink:0">{_icon_svg}</svg>
  <div>
    <div style="font-size:10px;font-weight:700;letter-spacing:1.4px;
                text-transform:uppercase;color:{status_clr};line-height:1;margin-bottom:3px">
        {status_lbl}
    </div>
    <div style="font-size:18px;font-weight:800;color:#0F1C2E;line-height:1.2">
        Revisión previa al análisis
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ─── ARCHIVO + ESTADO ────────────────────────────────────────────────
        st.markdown(f"""
<div style="display:flex;justify-content:space-between;align-items:center;
            background:#F4F8FF;border:1px solid #D4E2F4;border-radius:8px;
            padding:10px 14px;margin-bottom:14px">
  <div>
    <div style="font-size:9px;font-weight:700;text-transform:uppercase;
                letter-spacing:1px;color:#94A3B8;margin-bottom:2px">Archivo cargado</div>
    <div style="font-size:13px;font-weight:600;color:#1E293B">{fname}</div>
  </div>
  <div style="text-align:right">
    <div style="font-size:9px;font-weight:700;text-transform:uppercase;
                letter-spacing:1px;color:#94A3B8;margin-bottom:2px">Estado</div>
    <div style="font-size:12px;font-weight:700;
                color:{'#1E5FA8' if ok else '#D97706'}">
        {'Listo para analizar' if ok else 'Con advertencias'}
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ─── MÉTRICAS ────────────────────────────────────────────────────────
        st.markdown(f"""
<div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:10px;margin-bottom:14px">
  <div style="background:#F8FAFF;border:1px solid #DBEAFE;border-radius:8px;
              padding:10px 8px;text-align:center">
    <div style="font-size:8px;font-weight:700;text-transform:uppercase;
                letter-spacing:1px;color:#64748B;margin-bottom:4px">Total filas</div>
    <div style="font-size:24px;font-weight:900;color:#0F1C2E;line-height:1">{stats['total']:,}</div>
  </div>
  <div style="background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;
              padding:10px 8px;text-align:center">
    <div style="font-size:8px;font-weight:700;text-transform:uppercase;
                letter-spacing:1px;color:#64748B;margin-bottom:4px">Filas válidas</div>
    <div style="font-size:24px;font-weight:900;color:#1E5FA8;line-height:1">{stats['ok']:,}</div>
  </div>
  <div style="background:{'#FEF2F2' if stats['errors'] else '#F8FAFF'};
              border:1px solid {'#FECACA' if stats['errors'] else '#DBEAFE'};
              border-radius:8px;padding:10px 8px;text-align:center">
    <div style="font-size:8px;font-weight:700;text-transform:uppercase;
                letter-spacing:1px;color:#64748B;margin-bottom:4px">Con error</div>
    <div style="font-size:24px;font-weight:900;
                color:{'#DC2626' if stats['errors'] else '#94A3B8'};line-height:1">{stats['errors']:,}</div>
  </div>
  <div style="background:#F8FAFF;border:1px solid #DBEAFE;border-radius:8px;
              padding:10px 8px;text-align:center">
    <div style="font-size:8px;font-weight:700;text-transform:uppercase;
                letter-spacing:1px;color:#64748B;margin-bottom:4px">Cols. faltantes</div>
    <div style="font-size:{'12px' if stats['missing_cols'] else '24px'};font-weight:900;
                color:{'#D97706' if stats['missing_cols'] else '#94A3B8'};line-height:1;
                margin-top:{'2px' if stats['missing_cols'] else '0'}">{miss_val}</div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ─── SEPARADOR + TEXTO + BOTONES ─────────────────────────────────────
        st.markdown("""
<div style="border-top:1px solid #E8EDF3;padding-top:14px;margin-top:10px;margin-bottom:12px">
  <div style="font-size:13px;font-weight:600;color:#374151;text-align:center;line-height:1.4">
      Confirmá para procesar el archivo o Cancelá para volver al inicio.
  </div>
</div>
""", unsafe_allow_html=True)

        col_canc, col_cont = st.columns(2)
        with col_canc:
            if st.button("Cancelar", use_container_width=True, type="secondary", key="btn_cancelar"):
                for k in ["df", "fname", "confirmed", "stats", "tarifa_rows"]:
                    st.session_state.pop(k, None)
                st.rerun()

        with col_cont:
            if st.button("Continuar", use_container_width=True, type="primary", key="btn_continuar"):
                st.session_state["confirmed"] = True
                st.rerun()


import json
import os

TARIFAS_FILE = "tarifas.json"

def load_tarifas():
    if os.path.exists(TARIFAS_FILE):
        try:
            with open(TARIFAS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    return {}

def save_tarifas(data):
    try:
        with open(TARIFAS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except:
        pass

def main():
    if "confirmed" not in st.session_state:
        st.session_state["confirmed"] = False
    if "tarifa_map" not in st.session_state:
        st.session_state["tarifa_map"] = load_tarifas()

    with st.sidebar:
        st.markdown("### ⚙️ Ajustes")
        if st.button("Configurar Tarifas", use_container_width=True):
            st.session_state["show_settings"] = not st.session_state.get("show_settings", False)

    if st.session_state.get("show_settings", False):
        st.title("⚙️ Configuración de Tarifas")
        st.markdown("Configurá las tarifas por **Cliente** y **Localidad**. Usá **DEFAULT** en localidad si querés aplicar la tarifa a todas las demás zonas del cliente.")
        
        t_map = st.session_state.get("tarifa_map", {})
        flat_data = []
        for cli, d in t_map.items():
            if isinstance(d, dict):
                for loc, val in d.items():
                    flat_data.append({"Cliente": cli, "Localidad": loc, "Tarifa": val})
            else:
                flat_data.append({"Cliente": cli, "Localidad": "DEFAULT", "Tarifa": d})
                
        df_tar = pd.DataFrame(flat_data) if flat_data else pd.DataFrame(columns=["Cliente", "Localidad", "Tarifa"])
        
        # Obtener listas únicas combinando lo guardado en JSON y lo cargado
        saved_c = list(t_map.keys())
        saved_l = [loc for d in t_map.values() if isinstance(d, dict) for loc in d.keys()]
        df_tmp = st.session_state.get("df")
        curr_c, curr_l = [], []
        if df_tmp is not None:
            if "Cliente" in df_tmp.columns:
                curr_c = [str(x) for x in df_tmp["Cliente"].dropna().unique() if str(x).strip()]
            if "Localidad Destino" in df_tmp.columns:
                curr_l = [str(x) for x in df_tmp["Localidad Destino"].dropna().unique() if str(x).strip()]
                
        list_c = sorted(list(set(saved_c + curr_c)))
        list_l = sorted(list(set(saved_l + curr_l)))
        if "DEFAULT" not in list_l:
            list_l.insert(0, "DEFAULT")
            
        column_config = {
            "Cliente": st.column_config.SelectboxColumn("Cliente", options=list_c, required=True),
            "Localidad": st.column_config.SelectboxColumn("Localidad", options=list_l, required=True),
            "Tarifa": st.column_config.NumberColumn("Tarifa", min_value=0, step=1, required=True)
        }
        
        edited = st.data_editor(df_tar, num_rows="dynamic", use_container_width=True, column_config=column_config)
        
        if st.button("Guardar Tarifas", type="primary"):
            new_map = {}
            for _, r in edited.iterrows():
                c = str(r.get("Cliente", "")).strip()
                l = str(r.get("Localidad", "")).strip()
                t = r.get("Tarifa", 0)
                if c and pd.notna(t) and c.upper() != "NAN":
                    if not l or l.upper() == "NAN": l = "DEFAULT"
                    if c not in new_map: new_map[c] = {}
                    new_map[c][l] = int(t)
            st.session_state["tarifa_map"] = new_map
            save_tarifas(new_map)
            st.success("Tarifas guardadas exitosamente en tarifas.json.")
        
        if st.button("Volver al Inicio"):
            st.session_state["show_settings"] = False
            st.rerun()
        return

    if st.session_state.get("confirmed") and "df" in st.session_state:
        render_dashboard(st.session_state["df"], st.session_state["fname"])
        return

    if "df" in st.session_state and not st.session_state.get("confirmed"):
        render_popup(st.session_state["stats"], st.session_state["fname"])
        return

    uploaded = render_home()
    if not uploaded:
        return

    with st.spinner("Analizando archivo…"):
        try:
            df = load_file(uploaded)
        except Exception as e:
            st.error(f"No se pudo leer el archivo: {e}")
            return

    stats = validate_df(df)
    st.session_state["df"]        = df
    st.session_state["fname"]     = uploaded.name
    st.session_state["stats"]     = stats
    st.session_state["confirmed"] = False
    st.rerun()


if __name__ == '__main__':
    main()
